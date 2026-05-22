"""Ficha · Control Horario SaaS — RDL 8/2019 · ET Arts. 34/35/38 · RGPD — build:20260521"""

import streamlit as st

import re as _re
import textwrap as _textwrap
import os as _os

# ── App icon — load PNG relative to this file, fallback to emoji ──────────────
try:
    from PIL import Image as _PILImage
    _icon_path = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "icon.png")
    _APP_ICON = _PILImage.open(_icon_path)
except Exception:
    _APP_ICON = "🍸"
# ─────────────────────────────────────────────────────────────────────────────

# ── HTML render fix (definitive) ──────────────────────────────────────────────
# Two CommonMark bugs killed HTML rendering:
#   1. ≥4 leading spaces → code block  →  fix: textwrap.dedent().strip()
#   2. Blank line inside HTML block → block ends early  →  fix: re.sub blank lines
#
# Patch strategy: patch BOTH st.markdown (module-level) AND DeltaGenerator.markdown
# (container-level: col.markdown, tab.markdown, expander.markdown, etc.)
# The st.markdown patch is applied FIRST so it is always active.

def _clean_html(body: str) -> str:
    body = _textwrap.dedent(body).strip()
    body = _re.sub(r'\n[ \t]*\n', '\n', body)
    return body

# ① Patch st.markdown (module-level — catches all direct st.markdown() calls)
_st_md_orig = st.markdown
def _st_md_patched(body, **kwargs):
    if kwargs.get("unsafe_allow_html") and isinstance(body, str) and "<" in body:
        body = _clean_html(body)
    return _st_md_orig(body, **kwargs)
st.markdown = _st_md_patched

# ② Patch DeltaGenerator.markdown (catches col.markdown(), tab.markdown(), etc.)
try:
    from streamlit.delta_generator import DeltaGenerator as _DG
    _dg_md_orig = _DG.markdown
    def _dg_md_patched(self, body, **kwargs):
        if kwargs.get("unsafe_allow_html") and isinstance(body, str) and "<" in body:
            body = _clean_html(body)
        return _dg_md_orig(self, body, **kwargs)
    _DG.markdown = _dg_md_patched
except Exception:
    pass  # st.markdown already patched above — this is just extra coverage
# ─────────────────────────────────────────────────────────────────────────────

import pandas as pd

import plotly.express as px

import plotly.graph_objects as go

import calendar

from datetime import datetime, date, timedelta

from io import BytesIO

import base64 as _b64

import db



# ── Page config ───────────────────────────────────────────────────────────────

st.set_page_config(

    page_title="ODK Ficha",

    page_icon=_APP_ICON,

    layout="wide",

    initial_sidebar_state="expanded",

)



# ── iOS/Android home screen icon — triple approach ────────────────────────────
# iOS reads apple-touch-icon for "Add to Home Screen". Three layers:
#   1. st.markdown → <link> + <meta> tags injected into the page body
#      (iOS Safari scans the full document, not just <head>)
#   2. st.components iframe → window.parent.document.head patch (desktop/Android)
#   3. page_icon PIL image → browser tab favicon

def _inject_pwa_icons():
    icon_path = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "icon.png")
    if not _os.path.exists(icon_path):
        return
    with open(icon_path, "rb") as _f:
        _icon_b64 = _b64.b64encode(_f.read()).decode()
    _data_url = f"data:image/png;base64,{_icon_b64}"

    # Layer 1: direct tags via st.markdown (runs in main page context, not iframe)
    st.markdown(f"""
<link rel="apple-touch-icon" sizes="512x512" href="{_data_url}">
<link rel="apple-touch-icon-precomposed" sizes="512x512" href="{_data_url}">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-title" content="ODK Ficha">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<meta name="theme-color" content="#7C3AED">
<meta name="application-name" content="ODK Ficha">
""", unsafe_allow_html=True)

    # Layer 2: iframe → window.parent patch (extra coverage for non-iOS)
    st.components.v1.html(f"""
<script>
(function(){{
  var ico = '{_data_url}';
  function patch(doc) {{
    try {{
      var al = doc.querySelector('link[rel="apple-touch-icon"]');
      if (!al) {{ al = doc.createElement('link'); doc.head.appendChild(al); }}
      al.rel = 'apple-touch-icon'; al.sizes = '512x512'; al.href = ico;
      var lk = doc.querySelector('link[rel~="icon"]');
      if (!lk) {{ lk = doc.createElement('link'); lk.rel = 'icon'; doc.head.appendChild(lk); }}
      lk.type = 'image/png'; lk.href = ico;
      var tc = doc.querySelector('meta[name="theme-color"]');
      if (!tc) {{ tc = doc.createElement('meta'); doc.head.appendChild(tc); }}
      tc.name = 'theme-color'; tc.content = '#7C3AED';
    }} catch(e) {{}}
  }}
  patch(document);
  try {{ patch(window.parent.document); }} catch(e) {{}}
}})();
</script>
""", height=1)

_inject_pwa_icons()



# ── CSS ───────────────────────────────────────────────────────────────────────

STYLES = """

<style>

@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');



html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

.stApp { background: #f0f2f5; }

[data-testid="stSidebar"] { background: #1e1f2e !important; }

[data-testid="stSidebar"] * { color: #c9cde6 !important; }

[data-testid="stSidebar"] .stRadio label {

    padding: 8px 12px; border-radius: 8px; display:block;

    transition: background .2s; cursor:pointer;

}

[data-testid="stSidebar"] .stRadio label:hover { background: #2d2f45 !important; }



/* Cards */

.card {

    background: white; border-radius: 12px; padding: 20px;

    box-shadow: 0 1px 4px rgba(0,0,0,.08); margin-bottom: 16px;

}

.card-sm { background: white; border-radius: 10px; padding: 14px; box-shadow: 0 1px 3px rgba(0,0,0,.07); }



/* KPI metrics */

.kpi { text-align:center; padding:16px; background:white; border-radius:12px; box-shadow:0 1px 4px rgba(0,0,0,.08); }

.kpi-val { font-size:2.2rem; font-weight:700; color:#4f46e5; margin:0; }

.kpi-label { font-size:.85rem; color:#64748b; margin:0; }



/* Status badges */

.badge { display:inline-block; padding:3px 10px; border-radius:20px; font-size:.78rem; font-weight:600; }

.badge-verde    { background:#d1fae5; color:#065f46; }

.badge-amarillo { background:#fef3c7; color:#92400e; }

.badge-rojo     { background:#fee2e2; color:#991b1b; }

.badge-gris     { background:#f1f5f9; color:#475569; }

.badge-blue     { background:#dbeafe; color:#1e40af; }

.badge-purple   { background:#ede9fe; color:#5b21b6; }



/* Calendar */

.cal-wrapper { overflow-x:auto; }

.cal-table { width:100%; border-collapse:separate; border-spacing:4px; }

.cal-table th { text-align:center; font-size:.8rem; color:#94a3b8; font-weight:600; padding:6px; }

.cal-day {

    width:40px; height:40px; border-radius:8px; text-align:center;

    vertical-align:middle; font-size:.85rem; font-weight:500;

    cursor:pointer; transition:transform .1s, box-shadow .1s;

}

.cal-day:hover { transform:scale(1.1); box-shadow:0 2px 8px rgba(0,0,0,.15); }

.cal-verde    { background:#d1fae5; color:#065f46; }

.cal-amarillo { background:#fef9c3; color:#713f12; border:2px solid #fde047; }

.cal-rojo     { background:#fee2e2; color:#991b1b; border:2px solid #fca5a5; }

.cal-festivo  { background:#e0e7ff; color:#3730a3; }
.cal-festivo_auto { background:#fce7f3; color:#9d174d; }

.cal-fin_semana { background:#f8fafc; color:#cbd5e1; }

.cal-vacaciones { background:#fdf4ff; color:#7e22ce; }

.cal-futuro   { background:#f8fafc; color:#94a3b8; }

.cal-empty    { background:transparent; }

.cal-today    { outline:3px solid #4f46e5; }

.cal-selected { outline:3px solid #f59e0b; }



/* Fichaje big buttons */

.btn-ficha {

    display:block; width:100%; padding:18px; border-radius:14px;

    font-size:1.1rem; font-weight:700; border:none; cursor:pointer;

    transition:transform .1s, box-shadow .2s; text-align:center;

    box-shadow:0 4px 14px rgba(0,0,0,.12);

}

.btn-ficha:hover { transform:translateY(-2px); box-shadow:0 6px 20px rgba(0,0,0,.18); }

.btn-entrada { background:linear-gradient(135deg,#10b981,#059669); color:white; }

.btn-salida  { background:linear-gradient(135deg,#ef4444,#dc2626); color:white; }

.btn-pausa   { background:linear-gradient(135deg,#f59e0b,#d97706); color:white; }

.btn-fin_pausa { background:linear-gradient(135deg,#6366f1,#4f46e5); color:white; }



/* Status display */

.status-bar {

    background:white; border-radius:12px; padding:16px 20px;

    box-shadow:0 1px 4px rgba(0,0,0,.08);

    display:flex; align-items:center; gap:16px;

}

.status-dot { width:14px; height:14px; border-radius:50%; flex-shrink:0; }

.dot-trabajando { background:#10b981; animation:pulse 2s infinite; }

.dot-pausado    { background:#f59e0b; }

.dot-libre      { background:#94a3b8; }

.dot-completo   { background:#4f46e5; }

@keyframes pulse {

    0%,100% { box-shadow:0 0 0 0 rgba(16,185,129,.4); }

    50%      { box-shadow:0 0 0 8px rgba(16,185,129,0); }

}



/* Legal box */

.legal-box {

    background:#fefce8; border-left:4px solid #f59e0b;

    padding:12px 16px; border-radius:6px; font-size:.85rem;

}

.legal-box-red {

    background:#fff1f2; border-left:4px solid #f43f5e;

    padding:12px 16px; border-radius:6px; font-size:.85rem;

}



/* Login */

.login-card {

    max-width:420px; margin:80px auto; background:white;

    border-radius:20px; padding:40px; box-shadow:0 8px 40px rgba(0,0,0,.12);

}



/* Progress bar custom */

.prog-wrap { background:#f1f5f9; border-radius:20px; height:10px; overflow:hidden; margin:6px 0; }

.prog-fill  { height:100%; border-radius:20px; transition:width .3s; }

</style>

"""



# ── ODK Login CSS — dark luxury cocktail theme ────────────────────────────────

LOGIN_CSS = """

<style>
@import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800&family=Rubik:wght@300;400;500;600;700&display=swap');

/* ── Vibrant dark background — modern cocktail brand ── */
.stApp {
  background:
    radial-gradient(ellipse 65% 55% at 15% 85%, rgba(124,58,237,.22) 0%, transparent 55%),
    radial-gradient(ellipse 55% 50% at 85% 15%, rgba(244,63,94,.14) 0%, transparent 50%),
    radial-gradient(ellipse 40% 35% at 50% 50%, rgba(124,58,237,.06) 0%, transparent 60%),
    linear-gradient(145deg, #0a0812 0%, #0f0f23 55%, #0d0a18 100%) !important;
  min-height:100vh;
}

/* Hide Streamlit chrome */
[data-testid="stSidebar"]      { display:none !important; }
header[data-testid="stHeader"] { background:transparent !important; box-shadow:none !important; }
[data-testid="stToolbar"]      { display:none !important; }
[data-testid="stStatusWidget"] { display:none !important; }
[data-testid="stDeployButton"] { display:none !important; }
#MainMenu, footer              { display:none !important; visibility:hidden !important; }

.block-container { padding-top:4vh !important; padding-bottom:2rem !important; max-width:100% !important; }

/* ── ODK brand — bold, geometric, modern ── */
.odk-brand-wrap { text-align:center; padding:32px 20px 22px; }

.odk-logo-ring {
  display:inline-flex; align-items:center; justify-content:center;
  width:62px; height:62px; border-radius:18px;
  background:linear-gradient(135deg, #7C3AED, #F43F5E);
  margin-bottom:14px;
  box-shadow:0 0 32px rgba(124,58,237,.45), 0 8px 24px rgba(0,0,0,.3);
}

.odk-brand-name {
  font-family:'Outfit', sans-serif;
  font-weight:800; font-style:normal;
  font-size:2.8rem; letter-spacing:.18em; line-height:1;
  color:#fff;
  text-shadow:0 0 40px rgba(124,58,237,.5);
  display:block; margin-bottom:6px;
}

.odk-brand-sub {
  font-family:'Outfit', sans-serif;
  font-size:.72rem; font-weight:500;
  letter-spacing:.28em; text-transform:uppercase;
  color:rgba(226,232,240,.38);
  display:block; margin-top:4px;
}

.odk-accent-bar {
  width:40px; height:3px;
  background:linear-gradient(90deg, #7C3AED, #F43F5E);
  border-radius:2px;
  margin:12px auto 0;
}

/* ── Glass form card ── */
[data-testid="stForm"] {
  background:rgba(255,255,255,.055) !important;
  backdrop-filter:blur(20px) saturate(1.6) !important;
  -webkit-backdrop-filter:blur(20px) saturate(1.6) !important;
  border:1px solid rgba(124,58,237,.30) !important;
  border-radius:20px !important;
  padding:28px 26px 24px !important;
  box-shadow:
    0 0 0 1px rgba(124,58,237,.08),
    0 24px 64px rgba(0,0,0,.5),
    inset 0 1px 0 rgba(255,255,255,.08) !important;
}

/* ── Input labels ── */
.stTextInput > label {
  font-family:'Outfit', sans-serif !important;
  font-size:.72rem !important; font-weight:600 !important;
  letter-spacing:.14em !important; text-transform:uppercase !important;
  color:rgba(226,232,240,.60) !important;
}

/* ── Input fields ── */
.stTextInput input {
  background:rgba(255,255,255,.07) !important;
  border:1px solid rgba(124,58,237,.28) !important;
  border-radius:12px !important;
  color:#e2e8f0 !important;
  font-family:'Outfit', sans-serif !important;
  font-size:.95rem !important;
  transition:border-color .2s, box-shadow .2s !important;
}
.stTextInput input:focus {
  border-color:rgba(124,58,237,.75) !important;
  box-shadow:0 0 0 3px rgba(124,58,237,.16), 0 2px 12px rgba(124,58,237,.10) !important;
  outline:none !important;
}
.stTextInput input::placeholder { color:rgba(226,232,240,.22) !important; }

/* ── Submit button ── */
[data-testid="stFormSubmitButton"] > button {
  background:linear-gradient(135deg, #7C3AED 0%, #a855f7 50%, #F43F5E 100%) !important;
  border:none !important; border-radius:12px !important;
  color:#fff !important;
  font-family:'Outfit', sans-serif !important;
  font-weight:700 !important; font-size:.85rem !important;
  letter-spacing:.18em !important; text-transform:uppercase !important;
  padding:13px !important; width:100% !important;
  box-shadow:0 6px 28px rgba(124,58,237,.38), 0 2px 8px rgba(0,0,0,.25) !important;
  transition:all .22s ease !important;
  cursor:pointer !important;
}
[data-testid="stFormSubmitButton"] > button:hover {
  transform:translateY(-2px) !important;
  box-shadow:0 12px 36px rgba(124,58,237,.50), 0 4px 16px rgba(244,63,94,.20) !important;
}

/* ── Error alert ── */
[data-testid="stAlert"] {
  background:rgba(244,63,94,.10) !important;
  border:1px solid rgba(244,63,94,.28) !important;
  border-radius:12px !important; color:#fda4af !important;
}

/* ── Footer ── */
.odk-footer-txt {
  font-family:'Outfit', sans-serif;
  font-size:.72rem; letter-spacing:.04em;
  color:rgba(226,232,240,.28);
  text-align:center; margin-top:16px;
}
</style>

"""



# ── Helpers ───────────────────────────────────────────────────────────────────



def fmt_h(h: float) -> str:

    if h <= 0: return "0h 00m"

    return f"{int(h)}h {int((h%1)*60):02d}m"



def status_badge(status: str) -> str:

    MAP = {

        "verde":    ("badge-verde",    "✓ Completo"),

        "amarillo": ("badge-amarillo", "⚠ Incidencia"),

        "rojo":     ("badge-rojo",     "✗ Sin fichar"),

        "festivo":  ("badge-blue",     "Festivo"),

        "fin_semana": ("badge-gris",   "Fin de semana"),

        "vacaciones": ("badge-purple", "Vacaciones"),

        "futuro":   ("badge-gris",     "—"),

    }

    cls, label = MAP.get(status, ("badge-gris", status))

    return f'<span class="badge {cls}">{label}</span>'



def req_estado_badge(estado: str) -> str:

    M = {"pendiente": ("badge-amarillo","⏳ Pendiente"),

         "aprobada":  ("badge-verde",   "✓ Aprobada"),

         "denegada":  ("badge-rojo",    "✗ Denegada")}

    cls, label = M.get(estado, ("badge-gris", estado))

    return f'<span class="badge {cls}">{label}</span>'



COMUNIDADES_MAP = {

    "madrid": "Madrid", "cataluña": "Cataluña", "andalucia": "Andalucía",

    "valencia": "Valencia", "galicia": "Galicia", "pais_vasco": "País Vasco",

    "aragon": "Aragón", "castilla_leon": "Castilla y León",

    "castilla_la_mancha": "Castilla-La Mancha", "canarias": "Canarias",

    "baleares": "Baleares", "murcia": "Murcia", "navarra": "Navarra",

    "asturias": "Asturias", "cantabria": "Cantabria",

    "extremadura": "Extremadura", "la_rioja": "La Rioja",

}



MESES_ES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",

            "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

PROVINCIAS_ES = [
    "Álava","Albacete","Alicante","Almería","Asturias","Ávila",
    "Badajoz","Baleares","Barcelona","Burgos","Cáceres","Cádiz",
    "Cantabria","Castellón","Ciudad Real","Córdoba","A Coruña",
    "Cuenca","Girona","Granada","Guadalajara","Guipúzcoa","Huelva",
    "Huesca","Jaén","León","Lleida","La Rioja","Lugo","Madrid",
    "Málaga","Murcia","Navarra","Ourense","Palencia","Las Palmas",
    "Pontevedra","Salamanca","Santa Cruz de Tenerife","Segovia",
    "Sevilla","Soria","Tarragona","Teruel","Toledo","Valencia",
    "Valladolid","Vizcaya","Zamora","Zaragoza","Ceuta","Melilla",
]

PROVINCIA_COMUNIDAD = {
    "Álava": "pais_vasco", "Guipúzcoa": "pais_vasco", "Vizcaya": "pais_vasco",
    "Barcelona": "cataluña", "Girona": "cataluña", "Lleida": "cataluña",
    "Tarragona": "cataluña",
    "Las Palmas": "canarias", "Santa Cruz de Tenerife": "canarias",
    "Alicante": "valencia", "Castellón": "valencia", "Valencia": "valencia",
    "A Coruña": "galicia", "Lugo": "galicia", "Ourense": "galicia",
    "Pontevedra": "galicia",
    "Madrid": "madrid",
    "Sevilla": "andalucia", "Málaga": "andalucia", "Granada": "andalucia",
    "Córdoba": "andalucia", "Cádiz": "andalucia", "Jaén": "andalucia",
    "Almería": "andalucia", "Huelva": "andalucia",
    "Navarra": "navarra",
    "La Rioja": "la_rioja",
    "Asturias": "asturias",
    "Cantabria": "cantabria",
    "Murcia": "murcia",
    "Baleares": "baleares",
    "Zaragoza": "aragon", "Huesca": "aragon", "Teruel": "aragon",
    "Burgos": "castilla_leon", "León": "castilla_leon", "Salamanca": "castilla_leon",
    "Valladolid": "castilla_leon", "Zamora": "castilla_leon", "Palencia": "castilla_leon",
    "Ávila": "castilla_leon", "Segovia": "castilla_leon", "Soria": "castilla_leon",
    "Toledo": "castilla_la_mancha", "Ciudad Real": "castilla_la_mancha",
    "Albacete": "castilla_la_mancha", "Cuenca": "castilla_la_mancha",
    "Guadalajara": "castilla_la_mancha",
    "Badajoz": "extremadura", "Cáceres": "extremadura",
}



# ── Session state ─────────────────────────────────────────────────────────────



def ss(key, default=None):

    return st.session_state.get(key, default)



def logged_in() -> bool:

    return ss("user") is not None



def current_user() -> dict:

    return ss("user", {})



def is_role(*roles) -> bool:

    return current_user().get("role") in roles



# ── Login page ────────────────────────────────────────────────────────────────



def page_login():

    # Inject dark ODK luxury theme (overrides STYLES already injected by main())
    st.markdown(LOGIN_CSS, unsafe_allow_html=True)

    if st.session_state.get("login_blocked"):
        import time
        blocked_until = st.session_state.get("login_blocked_until", 0)
        if time.time() < blocked_until:
            remaining = int(blocked_until - time.time())
            st.error(f"🔒 Demasiados intentos fallidos. Espera {remaining} segundos.")
            return
        else:
            st.session_state["login_blocked"] = False
            st.session_state["login_attempts"] = 0

    _, col, _ = st.columns([1, 1.1, 1])

    with col:

        # ── ODK brand header — modern, bold, vibrant ─────────────────────────
        st.markdown("""
<div class="odk-brand-wrap">
  <div class="odk-logo-ring">
    <svg width="32" height="32" viewBox="0 0 32 32" fill="none" xmlns="http://www.w3.org/2000/svg">
      <path d="M16 4 C9.4 4 4 9.4 4 16 C4 22.6 9.4 28 16 28 C22.6 28 28 22.6 28 16 C28 9.4 22.6 4 16 4Z" stroke="rgba(255,255,255,0.4)" stroke-width="1.5" fill="none"/>
      <path d="M10 16 L16 10 L22 16 L19 16 L19 22 L13 22 L13 16Z" fill="white" opacity="0.9"/>
      <circle cx="23" cy="9" r="3" fill="rgba(255,255,255,0.7)"/>
    </svg>
  </div>
  <span class="odk-brand-name">ODK</span>
  <div class="odk-accent-bar"></div>
  <span class="odk-brand-sub">Ficha &middot; Control Horario</span>
</div>
""", unsafe_allow_html=True)

        # ── Login form ────────────────────────────────────────────────────────
        with st.form("login_form"):

            username = st.text_input("Usuario", placeholder="tu.usuario")

            password = st.text_input("Contraseña", type="password", placeholder="••••••••")

            submitted = st.form_submit_button("ACCEDER", use_container_width=True, type="primary")

        # ── Footer ────────────────────────────────────────────────────────────
        st.markdown(
            '<p class="odk-footer-txt">¿Problemas para acceder? Contacta con tu responsable.</p>',
            unsafe_allow_html=True
        )

        if submitted:

            user = db.authenticate(username.strip(), password)

            if user:

                st.session_state["login_attempts"] = 0
                st.session_state["login_blocked"] = False
                st.session_state["user"] = user

                db.audit(user["id"], "login", ip="0.0.0.0")

                st.rerun()

            else:

                st.session_state["login_attempts"] = st.session_state.get("login_attempts", 0) + 1
                if st.session_state["login_attempts"] >= 5:
                    import time
                    st.session_state["login_blocked"] = True
                    st.session_state["login_blocked_until"] = time.time() + 900  # 15 min
                st.error("Usuario o contraseña incorrectos.")



# ── Sidebar ───────────────────────────────────────────────────────────────────



def render_sidebar():

    user = current_user()

    unread = db.get_unread_count(user["id"])

    badge = f" 🔴" if unread > 0 else ""



    with st.sidebar:

        st.markdown(f"""

        <div style="padding:16px 0 8px">

            <div style="font-size:1.5rem;font-weight:700;color:#fff">⏱️ Ficha</div>

            <div style="font-size:.8rem;color:#7c7f9e;margin-top:2px">Control Horario SaaS</div>

        </div>

        <div style="background:#2d2f45;border-radius:10px;padding:12px;margin:8px 0 16px">

            <div style="font-weight:600;color:#e2e8f0">{user['nombre']} {user['apellidos']}</div>

            <div style="font-size:.78rem;color:#7c7f9e">{user['role'].upper()}</div>

        </div>

        """, unsafe_allow_html=True)



        # Admin items first so they don't get buried below personal items
        if is_role("admin"):

            nav_options = [
                "⏱️ Fichaje", "👤 Mi Perfil",
                "─── Admin ───",
                "👤 Usuarios", "📊 Dashboard Admin", "✅ Aprobar Solicitudes",
                "📅 Festivos", "📥 Exportaciones", "🔍 Auditoría",
                "─── Personal ───",
                "📅 Mi Calendario", "🏖️ Vacaciones", f"🔔 Notificaciones{badge}",
            ]

        elif is_role("manager"):

            nav_options = [
                "⏱️ Fichaje", "👤 Mi Perfil",
                "─── Equipo ───",
                "👥 Mi Equipo", "✅ Aprobar Solicitudes",
                "─── Personal ───",
                "📅 Mi Calendario", "🏖️ Vacaciones", f"🔔 Notificaciones{badge}",
            ]

        else:

            nav_options = ["⏱️ Fichaje", "👤 Mi Perfil", "📅 Mi Calendario", "🏖️ Vacaciones", f"🔔 Notificaciones{badge}"]



        page = st.radio("Navegación", nav_options, label_visibility="collapsed")



        st.divider()

        st.markdown("""

        <div style="font-size:.75rem;color:#4a4d6a;line-height:1.7">

            <div><b style="color:#7c7f9e">Normativa</b></div>

            <div>RDL 8/2019 — Registro diario</div>

            <div>Art.34 ET — 40h/sem · 9h/día</div>

            <div>Art.35 ET — Máx. 80h extra/año</div>

            <div>Art.38 ET — 30 días vacaciones</div>

        </div>

        """, unsafe_allow_html=True)



        if st.button("Cerrar sesión", use_container_width=True):

            st.session_state.clear()

            st.rerun()



    # Strip emoji prefix and notification badge; separators stay as-is (router will ignore them)
    if page.startswith("─"):
        return "Fichaje"  # fallback if separator somehow selected
    return page.split(" ", 1)[1].split("🔴")[0].strip() if " " in page else page




# ── Page: Mi Perfil ───────────────────────────────────────────────────────────



def page_perfil():

    user = current_user()

    st.title("👤 Mi Perfil")



    # ── Info card ─────────────────────────────────────────────────────────────

    try:

        from zoneinfo import ZoneInfo

        _tz_name = db.get_user_tz(user) if hasattr(db, 'get_user_tz') else "Europe/Madrid"

        zona_txt = "🌴 Canarias (Atlantic/Canary)" if _tz_name == "Atlantic/Canary" else "🇪🇸 Península (Europe/Madrid)"

    except Exception:

        zona_txt = "Europe/Madrid"



    c1, c2, c3 = st.columns(3)

    c1.metric("Nombre", f"{user.get('nombre','')} {user.get('apellidos','')}")

    c2.metric("Rol", user.get("role","").upper())

    c3.metric("Usuario", user.get("username",""))



    st.markdown("---")

    col1, col2 = st.columns(2)

    with col1:

        st.markdown("**📧 Email**")

        st.write(user.get("email") or "—")

        st.markdown("**📍 Provincia**")

        st.write(user.get("provincia") or "—")

        st.markdown("**🏙️ Localidad**")

        st.write(user.get("localidad") or "—")

    with col2:

        st.markdown("**🕐 Zona horaria (auto)**")

        st.write(zona_txt)

        st.markdown("**⏰ Horas semanales**")

        st.write(f"{user.get('horas_semanales', 40)}h")

        st.markdown("**🏖️ Días de vacaciones / año**")

        st.write(str(user.get("dias_vacaciones_anuales", 22)))



    # ── Vacation balance ──────────────────────────────────────────────────────

    bal = db.get_vacation_balance(user["id"], __import__("datetime").date.today().year)

    st.markdown("---")

    st.subheader("🏖️ Balance vacaciones")

    bc1, bc2, bc3, bc4 = st.columns(4)

    bc1.metric("Total", bal["total"])

    bc2.metric("Disfrutados", bal["used"])

    bc3.metric("Pendientes aprobación", bal["pending"])

    bc4.metric("Disponibles", bal["remaining"])



    # ── Change password ───────────────────────────────────────────────────────

    st.markdown("---")

    st.subheader("🔑 Cambiar contraseña")

    with st.form("cambiar_pw_form"):

        pw_actual  = st.text_input("Contraseña actual", type="password")

        pw_nueva   = st.text_input("Nueva contraseña", type="password")

        pw_confirm = st.text_input("Confirmar nueva contraseña", type="password")

        if st.form_submit_button("Cambiar contraseña"):

            if not pw_actual or not pw_nueva:

                st.error("Rellena todos los campos.")

            elif pw_nueva != pw_confirm:

                st.error("Las contraseñas nuevas no coinciden.")

            elif len(pw_nueva) < 8:

                st.error("La nueva contraseña debe tener al menos 8 caracteres.")

            else:

                verified = db.authenticate(user["username"], pw_actual)

                if not verified:

                    st.error("La contraseña actual no es correcta.")

                else:

                    db.update_user(user["id"], password=pw_nueva)

                    db.audit(user["id"], "cambiar_password", "users", user["id"])

                    st.success("✅ Contraseña actualizada correctamente.")



    # ── Admin note ────────────────────────────────────────────────────────────

    if not is_role("admin"):

        st.info("ℹ️ Para modificar tu provincia, localidad, email o datos laborales, contacta con el administrador.")

    else:

        st.info("✏️ Para editar todos los datos del perfil ve a **👤 Usuarios**.")


# ── Page: Fichaje ─────────────────────────────────────────────────────────────



def _offline_sync_js():
    """
    Inject offline-aware JS into the fichaje page.
    ─ Intercepts ENTRADA/SALIDA/PAUSA button clicks when navigator.onLine == false
    ─ Stores actions in localStorage with real timestamp
    ─ On reconnect (or page reload while online) encodes queue as ?sync=<b64>
      so the Python side can process and clear it.
    """
    st.components.v1.html("""
<script>
(function(){
  var KEY = 'odk_pending_fichajes';
  var TIPOS = {ENTRADA:'entrada', SALIDA:'salida', PAUSA:'pausa', REANUDAR:'fin_pausa'};

  // ── Toast notification ────────────────────────────────────────────────────
  function toast(msg, color) {
    var d = window.parent.document.createElement('div');
    d.innerText = msg;
    Object.assign(d.style, {
      position:'fixed', bottom:'24px', left:'50%', transform:'translateX(-50%)',
      background: color || '#7C3AED', color:'white', padding:'10px 22px',
      borderRadius:'24px', zIndex:'99999', fontSize:'14px', fontWeight:'600',
      boxShadow:'0 4px 20px rgba(0,0,0,.3)', transition:'opacity .5s'
    });
    window.parent.document.body.appendChild(d);
    setTimeout(function(){ d.style.opacity='0'; setTimeout(function(){ d.remove(); }, 600); }, 3000);
  }

  // ── Offline banner ────────────────────────────────────────────────────────
  var banner = null;
  function showBanner(show) {
    var doc = window.parent.document;
    if (show) {
      if (banner) return;
      banner = doc.createElement('div');
      banner.id = 'odk-offline-banner';
      banner.innerHTML = '📵 Sin conexión &nbsp;·&nbsp; Los fichajes se guardarán y sincronizarán al reconectar';
      Object.assign(banner.style, {
        position:'fixed', top:'0', left:'0', right:'0', zIndex:'99998',
        background:'linear-gradient(90deg,#7C3AED,#F43F5E)', color:'white',
        textAlign:'center', padding:'8px', fontSize:'13px', fontWeight:'600'
      });
      doc.body.prepend(banner);
    } else {
      var b = doc.getElementById('odk-offline-banner');
      if (b) b.remove(); banner = null;
    }
  }

  // ── On reconnect: reload with ?sync=... ──────────────────────────────────
  function syncIfPending() {
    var raw = localStorage.getItem(KEY);
    if (!raw) return;
    var queue = JSON.parse(raw);
    if (!queue.length) return;
    var enc = btoa(unescape(encodeURIComponent(JSON.stringify(queue))));
    var url = new URL(window.parent.location.href);
    url.searchParams.set('sync', enc);
    window.parent.location.href = url.toString();
  }

  window.addEventListener('online',  function(){ showBanner(false); toast('✅ Conexión restaurada — sincronizando...','#059669'); syncIfPending(); });
  window.addEventListener('offline', function(){ showBanner(true); toast('📵 Sin conexión — fichajes guardados localmente','#F43F5E'); });

  if (!navigator.onLine) showBanner(true);

  // ── Auto-sync on load if pending and online ──────────────────────────────
  if (navigator.onLine) syncIfPending();

  // ── Intercept button clicks when offline (capture phase) ─────────────────
  window.parent.document.addEventListener('click', function(e) {
    if (navigator.onLine) return;
    var btn = e.target.closest('[data-testid="stBaseButton-primary"] button, [data-testid="stBaseButton-secondary"] button, .stButton > button, [data-testid="stButton"] > button');
    if (!btn) return;
    var txt = btn.innerText.toUpperCase();
    var tipo = null;
    for (var k in TIPOS) { if (txt.indexOf(k) !== -1) { tipo = TIPOS[k]; break; } }
    if (!tipo) return;
    e.stopImmediatePropagation(); e.preventDefault();
    var queue = JSON.parse(localStorage.getItem(KEY) || '[]');
    queue.push({tipo: tipo, ts: new Date().toISOString()});
    localStorage.setItem(KEY, JSON.stringify(queue));
    toast('📵 ' + tipo.toUpperCase() + ' guardado — se sincronizará al reconectar', '#7C3AED');
  }, true);
})();
</script>
""", height=0)


def page_fichaje():

    user = current_user()

    # ── Offline sync: process any pending fichajes stored in localStorage ─────
    _sync_raw = st.query_params.get("sync", "")
    if _sync_raw:
        import json as _json
        try:
            _pending = _json.loads(_b64.b64decode(_sync_raw + "==").decode("utf-8", errors="replace"))
            _synced  = 0
            for _act in _pending:
                _tipo = _act.get("tipo")
                _ts   = _act.get("ts", "")
                if not _tipo:
                    continue
                try:
                    from zoneinfo import ZoneInfo
                    _dt  = datetime.fromisoformat(_ts).astimezone(ZoneInfo("Europe/Madrid"))
                    _dfe = _dt.date()
                    _hfe = _dt.strftime("%H:%M")
                except Exception:
                    _dfe = date.today()
                    _hfe = datetime.now().strftime("%H:%M")
                db.add_entry(user["id"], _dfe, _tipo, _hfe,
                             observaciones="[Offline — sync automático]",
                             is_manual=True, created_by=user["id"])
                db.audit(user["id"], f"fichaje_{_tipo}_offline", "time_entries", None,
                         {"fecha": str(_dfe), "hora": _hfe, "ts_original": _ts})
                _synced += 1
            if _synced:
                st.success(f"✅ {_synced} fichaje(s) offline sincronizados correctamente.")
            st.query_params.clear()
        except Exception as _ex:
            st.warning(f"⚠️ Error al sincronizar offline: {_ex}")
            st.query_params.clear()

    # ── Inject offline JS support ─────────────────────────────────────────────
    _offline_sync_js()

    # ── Timezone: auto from provincia + manual override ──────────────────────
    # IMPORTANT: read the widget key 'tz_sel' directly here (Streamlit updates
    # widget keys in session_state BEFORE the script reruns, so reading it at
    # the top gives the correct value immediately after the user changes it).

    try:

        from zoneinfo import ZoneInfo

        _auto_tz = db.get_user_tz(user) if hasattr(db, 'get_user_tz') else "Europe/Madrid"

        _tz_sel_now = st.session_state.get("tz_sel", "Auto")

        if _tz_sel_now == "🌴 Canarias":

            _tz_name = "Atlantic/Canary"

        elif _tz_sel_now == "🇪🇸 Península":

            _tz_name = "Europe/Madrid"

        else:

            _tz_name = _auto_tz

        _tz  = ZoneInfo(_tz_name)

        _now = datetime.now(_tz)

    except Exception:

        _auto_tz = "Europe/Madrid"

        _tz_name = "Europe/Madrid"

        _now = datetime.now()

    _is_canary = _tz_name == "Atlantic/Canary"

    zona_label = "🌴 Canarias" if _is_canary else "🇪🇸 Península"



    hoy   = _now.date()

    ahora = _now.strftime("%H:%M")



    entries = db.get_day_entries(user["id"], hoy)

    state   = db.get_fichaje_state(user["id"], hoy)

    worked  = db.calc_worked_hours(entries)

    last_hora = entries[-1]['hora'][:5] if entries else "--:--"



    # ── Extra CSS ──────────────────────────────────────────────────────────────

    st.markdown("""<style>

    .ficha-hero {

        background: linear-gradient(135deg,#1e1f2e 0%,#2d2f45 100%);

        border-radius:20px; padding:28px 32px; margin-bottom:20px;

        display:flex; align-items:center; justify-content:space-between;

        box-shadow:0 8px 32px rgba(0,0,0,.2);

    }

    .ficha-time  { font-size:3.2rem; font-weight:800; color:white; letter-spacing:-2px; line-height:1; }

    .ficha-date  { font-size:.9rem; color:#a0aec0; margin-top:4px; }

    .status-pill {

        display:inline-flex; align-items:center; gap:8px;

        padding:10px 20px; border-radius:50px; font-weight:600; font-size:.95rem;

        box-shadow:0 4px 12px rgba(0,0,0,.2);

    }

    .pill-libre      { background:rgba(148,163,184,.2); color:#94a3b8; }

    .pill-trabajando { background:rgba(16,185,129,.2);  color:#34d399; border:1px solid rgba(16,185,129,.35); }

    .pill-pausado    { background:rgba(245,158,11,.2);  color:#fbbf24; border:1px solid rgba(245,158,11,.35); }

    .pill-completo   { background:rgba(99,102,241,.2);  color:#a5b4fc; border:1px solid rgba(99,102,241,.35); }

    .tl-row {

        display:flex; align-items:center; gap:12px;

        padding:10px 0; border-bottom:1px solid #f1f5f9;

    }

    .tl-icon {

        width:36px; height:36px; border-radius:50%; display:flex;

        align-items:center; justify-content:center; font-size:1rem; flex-shrink:0;

    }

    </style>""", unsafe_allow_html=True)



    # ── Hero card ──────────────────────────────────────────────────────────────

    pill_cls = {"entrada":"pill-trabajando","fin_pausa":"pill-trabajando",

                "pausa":"pill-pausado","salida":"pill-completo"}.get(state,"pill-libre")

    pill_txt = {

        None:       "● Sin fichar hoy",

        "entrada":  f"● Trabajando desde {last_hora}",

        "pausa":    f"● En pausa desde {last_hora}",

        "fin_pausa":f"● Activo tras pausa ({last_hora})",

        "salida":   f"✓ Jornada finalizada",

    }.get(state, "—")



    pct    = min(int(worked / 8 * 100), 100)

    bar_color = "#ef4444" if worked > 9 else "#10b981" if worked >= 8 else "#6366f1"

    overtime_badge = ('<span style="background:#fca5a5;color:#991b1b;padding:2px 8px;'

                      'border-radius:10px;font-size:.72rem;font-weight:700;margin-left:8px">⚠ +9h</span>'

                      if worked > 9 else "")



    st.markdown(f"""

    <div class="ficha-hero">

      <div>

        <div class="ficha-time">{ahora}</div>

        <div class="ficha-date">{_now.strftime('%A, %d de %B de %Y').capitalize()} &nbsp;·&nbsp; {zona_label}</div>

        <div style="margin-top:18px">

          <div style="font-size:.78rem;color:#94a3b8;margin-bottom:5px">

            Horas trabajadas hoy {overtime_badge}

          </div>

          <div style="background:rgba(255,255,255,.12);border-radius:20px;height:8px;width:220px;overflow:hidden">

            <div style="background:{bar_color};height:100%;width:{pct}%;border-radius:20px;transition:width .4s"></div>

          </div>

          <div style="font-size:.82rem;color:#c9cde6;margin-top:5px"><b>{fmt_h(worked)}</b> de 8h objetivo</div>

        </div>

      </div>

      <div style="text-align:right">

        <div class="status-pill {pill_cls}">{pill_txt}</div>

        <div style="margin-top:10px;font-size:.78rem;color:#6b7280">

          Usuario: <b style="color:#a0aec0">{user.get('nombre','')}</b>

        </div>

      </div>

    </div>

    """, unsafe_allow_html=True)



    # ── Config row (jornada + hora + zona horaria) ──────────────────────────────

    col_j, col_h, col_tz, col_obs = st.columns([2, 1, 1.5, 3])

    with col_j:

        tipo_jornada = st.selectbox(

            "Modalidad",

            ["🏢 Presencial", "🏠 Teletrabajo", "🔄 Mixta", "🚗 Desplazamiento"],

            key="tipo_jornada",

        )

    with col_h:

        hora_manual_cb = st.checkbox("⏱ Hora manual", key="hora_manual_cb")

    with col_tz:

        _tz_opts = ["Auto", "🌴 Canarias", "🇪🇸 Península"]

        _tz_default = 1 if _is_canary and _tz_name != _auto_tz else 2 if not _is_canary and _tz_name != _auto_tz else 0

        st.selectbox("Zona horaria", _tz_opts, index=_tz_default, key="tz_sel", help="Auto detecta desde tu perfil. Cambia para sobreescribir.")

    with col_obs:

        obs = st.text_input("Observación (opcional)",

                            placeholder="Ej: Reunión con cliente…",

                            key="obs_fichaje",

                            label_visibility="collapsed")



    if hora_manual_cb:

        hora_input = st.time_input(

            "Hora del fichaje",

            value=_now.time().replace(second=0, microsecond=0),

            step=60,

            key="hora_input_manual",

        )

        hora_str = hora_input.strftime("%H:%M")

    else:

        hora_str = ahora



    # ── Action buttons ─────────────────────────────────────────────────────────

    can_entrada   = state is None

    can_pausa     = state in ("entrada", "fin_pausa")

    can_fin_pausa = state == "pausa"

    can_salida    = state in ("entrada", "fin_pausa")



    def fichar(tipo):

        eid = db.add_entry(user["id"], hoy, tipo, hora_str,

                           observaciones=obs,

                           is_manual=hora_manual_cb,

                           created_by=user["id"])

        db.audit(user["id"], f"fichaje_{tipo}", "time_entries", eid,

                 {"fecha": str(hoy), "hora": hora_str})

        st.rerun()



    active = []

    if can_entrada:   active.append(("entrada",   "🟢 ENTRADA",   "primary"))

    if can_pausa:     active.append(("pausa",     "🟡 PAUSA",     "secondary"))

    if can_fin_pausa: active.append(("fin_pausa", "🟣 REANUDAR",  "primary"))

    if can_salida:    active.append(("salida",    "🔴 SALIDA",    "secondary"))



    if active:

        btn_cols = st.columns(len(active))

        for idx, (tipo, label, btype) in enumerate(active):

            with btn_cols[idx]:

                if st.button(label, use_container_width=True,

                             type="primary" if btype == "primary" else "secondary",

                             key=f"btn_{tipo}"):

                    worked_at_exit = db.calc_worked_hours(entries)

                    fichar(tipo)

                    if tipo == "salida" and worked_at_exit > 9:

                        db.create_incident(user["id"], hoy, "jornada_excesiva",

                                           f"Jornada de {worked_at_exit:.2f}h supera 9h (Art.34.3 ET)")

    elif state == "salida":

        st.success("✅ ¡Jornada completada! Buen descanso.")



    st.divider()



    # ── Timeline + stats ───────────────────────────────────────────────────────

    col_tl, col_stats = st.columns([3, 2])



    with col_tl:

        st.markdown(f"##### 📋 Registros de hoy — {hoy.strftime('%d/%m/%Y')}")

        if entries:

            TIPO_COLORS = {

                "entrada":   ("#d1fae5", "#065f46", "🟢"),

                "salida":    ("#fee2e2", "#991b1b", "🔴"),

                "pausa":     ("#fef3c7", "#92400e", "🟡"),

                "fin_pausa": ("#ede9fe", "#5b21b6", "🟣"),

            }

            for e in entries:

                bg, fg, ico = TIPO_COLORS.get(e['tipo'], ("#f1f5f9","#475569","⚪"))

                extra = []

                if e["is_manual"]: extra.append("✏️ Manual")

                if e.get("observaciones"): extra.append(e["observaciones"])

                c_ico, c_txt, c_del = st.columns([0.4, 4, 0.6])

                with c_ico:

                    st.markdown(

                        f'<div class="tl-icon" style="background:{bg};color:{fg};margin-top:6px">{ico}</div>',

                        unsafe_allow_html=True)

                with c_txt:

                    label = f"**{e['hora'][:5]}** — {e['tipo'].replace('_',' ').title()}"

                    if extra:

                        label += f"  \n*{' · '.join(extra)}*"

                    st.markdown(label)

                with c_del:

                    if st.button("🗑️", key=f"del_{e['id']}",

                                 help="Eliminar este registro",

                                 use_container_width=True):

                        db.soft_delete_entry(e["id"], user["id"])

                        db.audit(user["id"], "delete_entry", "time_entries", e["id"],

                                 {"fecha": str(hoy), "tipo": e["tipo"], "hora": e["hora"]})

                        st.rerun()

        else:

            st.markdown("""

            <div style="text-align:center;padding:40px 20px;color:#94a3b8">

              <div style="font-size:2.5rem;margin-bottom:8px">⏰</div>

              <div>Sin registros hoy.<br>Pulsa <b>ENTRADA</b> para comenzar.</div>

            </div>""", unsafe_allow_html=True)



    with col_stats:

        st.markdown("##### 📊 Resumen del día")

        st.markdown(f"""

        <div class="card" style="text-align:center;padding:24px">

          <div style="font-size:2.8rem;font-weight:800;color:#4f46e5;line-height:1">{fmt_h(worked)}</div>

          <div style="font-size:.82rem;color:#64748b;margin:6px 0 14px">trabajadas hoy</div>

          <div style="background:#f1f5f9;border-radius:20px;height:10px;overflow:hidden">

            <div style="background:{bar_color};height:100%;width:{pct}%;border-radius:20px"></div>

          </div>

          <div style="display:flex;justify-content:space-between;font-size:.72rem;color:#94a3b8;margin-top:4px">

            <span>0h</span><span>8h</span>

          </div>

        </div>""", unsafe_allow_html=True)

        if worked > 9:

            st.error("⚠️ Supera 9h diarias (Art. 34.3 ET)")

        elif worked > 8:

            st.warning(f"Jornada extendida: {fmt_h(worked)}")

        elif worked >= 8 and state == "salida":

            st.success("✅ Jornada completa")



    # ── Corrección manual ──────────────────────────────────────────────────────

    with st.expander("✏️ Corregir o añadir fichaje en otra fecha"):

        with st.form("manual_form"):

            mc1, mc2, mc3 = st.columns(3)

            with mc1:

                fecha_m = st.date_input("Fecha", value=hoy, max_value=hoy)

            with mc2:

                tipo_m = st.selectbox("Tipo", ["entrada", "salida", "pausa", "fin_pausa"])

            with mc3:

                hora_m = st.time_input(

                    "Hora",

                    value=_now.time().replace(second=0, microsecond=0),

                    step=60,

                )

            es_hoy = (fecha_m == hoy)

            obs_m = st.text_input(

                "Observación (opcional)" if es_hoy else "Motivo / justificación *",

                placeholder="" if es_hoy else "Obligatorio para fechas pasadas",

            )

            reemplazar = st.checkbox(

                "Reemplazar registro existente del mismo tipo",

                value=True,

                help="Si ya hay una Entrada/Salida de ese tipo en esa fecha, se elimina la antigua y se guarda la nueva hora.",

            )

            if st.form_submit_button("💾 Guardar fichaje", use_container_width=True):

                if not es_hoy and not obs_m.strip():

                    st.error("El motivo es obligatorio para fechas pasadas.")

                else:

                    if reemplazar:

                        # Soft-delete ALL existing entries of the same tipo on that date

                        existing = db.get_day_entries(user["id"], fecha_m)

                        for old_e in existing:

                            if old_e["tipo"] == tipo_m:

                                db.soft_delete_entry(old_e["id"], user["id"])

                                db.audit(user["id"], "delete_entry_replaced",

                                         "time_entries", old_e["id"],

                                         {"fecha": str(fecha_m), "tipo": tipo_m,

                                          "hora_old": old_e["hora"]})

                    db.add_entry(user["id"], fecha_m, tipo_m,

                                 hora_m.strftime("%H:%M"),

                                 observaciones=obs_m, is_manual=True,

                                 created_by=user["id"])

                    db.audit(user["id"], "fichaje_manual", "time_entries",

                             datos={"fecha": str(fecha_m), "tipo": tipo_m})

                    st.success("✅ Fichaje guardado correctamente.")

                    st.rerun()



    # ── Legal notice ───────────────────────────────────────────────────────────

    st.markdown("""

    <div class="legal-box" style="margin-top:16px">

        📋 <b>RDL 8/2019</b>: Registro de jornada obligatorio. Datos conservados

        <b>4 años</b> y disponibles para la Inspección de Trabajo (Art. 34.9 ET).

    </div>

    """, unsafe_allow_html=True)



# ── Page: Calendario ──────────────────────────────────────────────────────────



def render_calendar(user_id, año, mes, comunidad="madrid"):

    """Returns HTML string for a month calendar with colored days."""

    cal = calendar.monthcalendar(año, mes)

    hoy = date.today()

    sel = ss("cal_selected_day")



    headers = "".join(f"<th>{d}</th>" for d in ["L","M","X","J","V","S","D"])

    rows = ""

    for week in cal:

        row = "<tr>"

        for i, day in enumerate(week):

            if day == 0:

                row += '<td class="cal-day cal-empty"></td>'

            else:

                d = date(año, mes, day)

                status = db.get_day_status(user_id, d, comunidad)

                today_cls = " cal-today" if d == hoy else ""

                sel_cls = " cal-selected" if sel == str(d) else ""

                row += (f'<td class="cal-day cal-{status}{today_cls}{sel_cls}" '

                        f'title="{d}">{day}</td>')

        row += "</tr>"

        rows += row



    return f"""

    <div class="cal-wrapper">

    <table class="cal-table">

    <thead><tr>{headers}</tr></thead>

    <tbody>{rows}</tbody>

    </table>

    </div>

    """



def page_calendario():

    user = current_user()

    hoy = date.today()



    st.title("📅 Mi Calendario")



    if "cal_year" not in st.session_state:

        st.session_state["cal_year"] = hoy.year

        st.session_state["cal_month"] = hoy.month



    año = st.session_state["cal_year"]

    mes = st.session_state["cal_month"]



    c_left, c_mid, c_right = st.columns([1, 3, 1])

    with c_left:

        if st.button("← Anterior"):

            if mes == 1:

                st.session_state["cal_month"] = 12

                st.session_state["cal_year"] = año - 1

            else:

                st.session_state["cal_month"] -= 1

            st.rerun()

    with c_mid:

        st.markdown(f"<h3 style='text-align:center;margin:0'>{MESES_ES[mes-1]} {año}</h3>",

                    unsafe_allow_html=True)

    with c_right:

        if st.button("Siguiente →"):

            if mes == 12:

                st.session_state["cal_month"] = 1

                st.session_state["cal_year"] = año + 1

            else:

                st.session_state["cal_month"] += 1

            st.rerun()



    comunidad = user.get("comunidad_autonoma", "madrid")

    cal_html = render_calendar(user["id"], año, mes, comunidad)



    st.markdown(f'<div class="card">{cal_html}</div>', unsafe_allow_html=True)



    # Leyenda separada (Streamlit 1.36+ filtra HTML tras </table>)

    st.markdown(

        "🟩 Completo &nbsp; 🟨 Incidencia &nbsp; 🟥 Sin fichar "

        "&nbsp; 🟦 Festivo &nbsp; 🟪 Vacaciones",

        unsafe_allow_html=True,

    )



    # Day selector for details

    st.subheader("Detalle del día")

    col1, col2 = st.columns([1, 2])

    with col1:

        sel_day = st.date_input("Selecciona un día", value=hoy,

                                min_value=date(año, mes, 1),

                                max_value=date(año, mes, calendar.monthrange(año, mes)[1]))

    with col2:

        status = db.get_day_status(user["id"], sel_day, comunidad)

        entries = db.get_day_entries(user["id"], sel_day)

        worked = db.calc_worked_hours(entries)



        st.markdown(f"**Estado:** {status_badge(status)}", unsafe_allow_html=True)

        st.markdown(f"**Horas trabajadas:** {fmt_h(worked)}")



        if entries:

            tipo_icons = {"entrada":"🟢","salida":"🔴","pausa":"🟡","fin_pausa":"🟣"}

            for e in entries:

                manual = " ✏️ *manual*" if e["is_manual"] else ""

                st.markdown(f"- {tipo_icons.get(e['tipo'],'⚪')} **{e['hora'][:5]}** {e['tipo'].replace('_',' ').title()}{manual}")

        else:

            if sel_day <= hoy and sel_day.weekday() < 5 and status not in ("festivo","vacaciones"):

                st.warning("Sin registros — día laboral sin fichar.")



        # Incidencias del día

        incidents = [i for i in db.get_user_incidents(user["id"])

                     if i["fecha"] == str(sel_day)]

        if incidents:

            for inc in incidents:

                st.markdown(f"⚠️ **Incidencia**: {inc['tipo'].replace('_',' ').title()} — {inc['descripcion']}")



    # Monthly validation

    st.subheader(f"Validación del mes — {MESES_ES[mes-1]} {año}")

    val = db.get_monthly_validation(user["id"], año, mes)

    if val and val["validado"]:

        st.success(f"✅ Mes validado el {val['fecha_validacion'][:10]}")

    else:

        if mes <= hoy.month and año <= hoy.year:

            if st.button("✅ Confirmar que mis registros de este mes son correctos"):

                db.set_monthly_validation(user["id"], año, mes)

                db.audit(user["id"], "validacion_mensual", datos={"año": año, "mes": mes})

                st.success("Mes validado correctamente. Registro guardado con timestamp.")

                st.rerun()

        else:

            st.info("Podrás validar este mes cuando llegue la fecha.")



    # Stats for the month

    f_ini = date(año, mes, 1)

    last_day = calendar.monthrange(año, mes)[1]

    f_fin = date(año, mes, last_day)

    entries_mes = db.get_entries_range(user["id"], f_ini, f_fin)



    # Build full month grid (all working days, even with 0 hours)

    hoy_local = date.today()

    all_days = [date(año, mes, d) for d in range(1, last_day + 1)

                if date(año, mes, d).weekday() < 5               # skip weekends

                and date(año, mes, d) <= hoy_local]              # skip future



    if entries_mes:

        df = pd.DataFrame(entries_mes)

        worked_by_day = (

            df.groupby("fecha")

            .apply(lambda g: db.calc_worked_hours(g.to_dict("records")))

            .to_dict()

        )

    else:

        worked_by_day = {}



    if all_days:

        days_str  = [d.strftime("%d %b") for d in all_days]

        hours     = [worked_by_day.get(str(d), 0.0) for d in all_days]

        bar_colors = [

            "#10b981" if h >= 8 else

            "#f59e0b" if h > 0  else

            "#ef4444"

            for h in hours

        ]



        fig = go.Figure(go.Bar(

            x=days_str,

            y=hours,

            marker_color=bar_colors,

            text=[fmt_h(h) if h > 0 else "" for h in hours],

            textposition="outside",

            hovertemplate="<b>%{x}</b><br>%{y:.2f}h trabajadas<extra></extra>",

        ))

        fig.add_hline(y=8, line_dash="dash", line_color="#6366f1",

                      annotation_text="8h objetivo", annotation_position="right")

        fig.add_hline(y=9, line_dash="dash", line_color="#ef4444",

                      annotation_text="Máx. 9h", annotation_position="right")

        fig.update_layout(

            title=dict(text=f"Horas trabajadas — {MESES_ES[mes-1]} {año}",

                       font=dict(size=15)),

            xaxis=dict(title="", tickangle=-45, tickfont=dict(size=11)),

            yaxis=dict(title="Horas", range=[0, max(max(hours, default=0) * 1.2, 10)]),

            plot_bgcolor="white",

            paper_bgcolor="white",

            height=320,

            margin=dict(t=50, b=60, l=40, r=80),

            showlegend=False,

        )

        st.plotly_chart(fig, use_container_width=True)

        # Mini legend

        st.caption("🟢 ≥ 8h completo &nbsp; 🟡 parcial &nbsp; 🔴 sin registros")



# ── Page: Vacaciones ──────────────────────────────────────────────────────────



def page_vacaciones():

    user = current_user()

    hoy = date.today()

    año = hoy.year



    st.title("🏖️ Vacaciones y Ausencias")



    bal = db.get_vacation_balance(user["id"], año)



    # Balance cards

    c1, c2, c3, c4 = st.columns(4)

    c1.metric("📅 Días totales", bal["total"])

    c2.metric("✅ Disfrutados", bal["used"])

    c3.metric("⏳ Pendientes aprobación", bal["pending"])

    c4.metric("🟢 Disponibles", bal["remaining"])



    pct = int((bal["used"] / bal["total"]) * 100) if bal["total"] > 0 else 0

    st.markdown(f"""

    <div style="margin:8px 0 20px">

        <div style="font-size:.85rem;color:#64748b;margin-bottom:4px">

            Consumo de vacaciones {año} — {pct}% utilizado

        </div>

        <div class="prog-wrap">

            <div class="prog-fill" style="width:{pct}%;background:{'#ef4444' if pct>90 else '#4f46e5'}"></div>

        </div>

    </div>

    """, unsafe_allow_html=True)



    tab_sol, tab_nueva = st.tabs(["📋 Mis Solicitudes", "➕ Nueva Solicitud"])



    with tab_sol:

        reqs = db.get_user_requests(user["id"])

        if not reqs:

            st.info("No tienes solicitudes registradas.")

        else:

            for r in reqs:

                with st.expander(

                    f"{r['tipo'].replace('_',' ').title()} · "

                    f"{r['fecha_inicio']} → {r['fecha_fin']} "

                    f"({r['dias_laborables']} días laborables)"

                ):

                    st.markdown(f"**Estado:** {req_estado_badge(r['estado'])}", unsafe_allow_html=True)

                    if r.get("manager_nombre"):

                        st.caption(f"Gestionado por: {r['manager_nombre']}")

                    if r["comentario_empleado"]:

                        st.write(f"Tu comentario: {r['comentario_empleado']}")

                    if r["comentario_manager"]:

                        st.write(f"Respuesta manager: {r['comentario_manager']}")

                    st.caption(f"Solicitado: {r['fecha_solicitud'][:10]}")

                    if r.get("fecha_resolucion"):

                        st.caption(f"Resuelto: {r['fecha_resolucion'][:10]}")



    with tab_nueva:

        TIPOS = ["vacaciones", "asuntos_propios", "descanso_compensatorio",

                 "baja_medica", "permiso_retribuido", "maternidad_paternidad",

                 "lactancia", "otros"]

        comunidad = user.get("comunidad_autonoma", "madrid")

        with st.form("nueva_solicitud"):

            tipo = st.selectbox("Tipo de ausencia", TIPOS,

                                format_func=lambda x: x.replace("_", " ").title())

            c1, c2 = st.columns(2)

            f_ini = c1.date_input("Fecha inicio", value=hoy + timedelta(days=1))

            f_fin = c2.date_input("Fecha fin", value=hoy + timedelta(days=1))

            comentario = st.text_area("Comentario / motivo (opcional)")



            if f_fin >= f_ini:

                dl = db.dias_laborables(f_ini, f_fin, comunidad)

                dn = (f_fin - f_ini).days + 1

                st.info(f"📅 **{dn} días naturales · {dl} días laborables**")

                if tipo == "vacaciones" and dl > bal["remaining"]:

                    st.warning(f"⚠️ Solicitas {dl} días pero solo tienes {bal['remaining']} disponibles.")



            submitted = st.form_submit_button("📤 Enviar solicitud", type="primary")

            if submitted:

                if f_fin < f_ini:

                    st.error("La fecha fin no puede ser anterior al inicio.")

                else:

                    rid = db.create_vac_request(user["id"], f_ini, f_fin, tipo,

                                                comentario, comunidad)

                    # notify manager

                    mgr_id = user.get("manager_id")

                    if mgr_id:

                        db.add_notification(mgr_id, "solicitud_vacaciones",

                            f"Nueva solicitud de {user['nombre']} {user['apellidos']}",

                            f"{tipo.replace('_',' ').title()}: {f_ini} → {f_fin} ({db.dias_laborables(f_ini,f_fin,comunidad)} días)")

                    db.audit(user["id"], "solicitud_vacaciones", "vacation_requests", rid)

                    st.success("✅ Solicitud enviada. Tu manager recibirá una notificación.")

                    st.rerun()



        st.markdown("""<div class="legal-box">

        <b>Marco legal:</b> Vacaciones mínimas 30 días naturales/año (Art. 38 ET) ·

        Permisos retribuidos: matrimonio 15d, nacimiento, fallecimiento (Art. 37 ET) ·

        No compensables en metálico salvo extinción de contrato.

        </div>""", unsafe_allow_html=True)



# ── Page: Notificaciones ──────────────────────────────────────────────────────



def page_notificaciones():

    user = current_user()

    st.title("🔔 Notificaciones e Incidencias")



    col_notif, col_inc = st.columns([1, 1])



    with col_notif:

        st.subheader("Notificaciones")

        if st.button("Marcar todas como leídas"):

            db.mark_all_read(user["id"])

            st.rerun()



        notifs = db.get_notifications(user["id"])

        if not notifs:

            st.info("Sin notificaciones.")

        else:

            for n in notifs:

                icon = "🔵" if not n["leido"] else "⚪"

                with st.container():

                    st.markdown(f"""

                    <div class="card-sm" style="margin-bottom:8px;{'border-left:3px solid #4f46e5' if not n['leido'] else ''}">

                        <div style="font-weight:{'600' if not n['leido'] else '400'};color:#1e293b">

                            {icon} {n['titulo']}

                        </div>

                        <div style="font-size:.85rem;color:#64748b">{n['mensaje']}</div>

                        <div style="font-size:.75rem;color:#94a3b8">{n['created_at'][:16]}</div>

                    </div>

                    """, unsafe_allow_html=True)

                    if not n["leido"]:

                        if st.button("Marcar leída", key=f"nr_{n['id']}"):

                            db.mark_read(n["id"])

                            st.rerun()



    with col_inc:

        st.subheader("Mis Incidencias")

        incidents = db.get_user_incidents(user["id"])



        # Auto-detect new incidents for past 7 days

        for i in range(1, 8):

            d = date.today() - timedelta(days=i)

            if d.weekday() >= 5: continue

            status = db.get_day_status(user["id"], d, user.get("comunidad_autonoma","madrid"))

            if status == "rojo":

                db.create_incident(user["id"], d, "sin_registro",

                                   f"Día laboral {d} sin ningún fichaje")

            elif status == "amarillo":

                db.create_incident(user["id"], d, "dia_incompleto",

                                   f"Día {d} con registros incompletos")



        incidents = db.get_user_incidents(user["id"])



        if not incidents:

            st.success("✅ Sin incidencias. Cumplimiento perfecto.")

        else:

            for inc in incidents:

                color = {"pendiente": "#fef3c7", "resuelto": "#d1fae5", "justificado": "#dbeafe"}.get(inc["estado"], "#f1f5f9")

                st.markdown(f"""

                <div style="background:{color};border-radius:8px;padding:12px;margin-bottom:8px">

                    <div style="font-weight:600">{inc['tipo'].replace('_',' ').title()} — {inc['fecha']}</div>

                    <div style="font-size:.85rem">{inc['descripcion']}</div>

                    <div style="font-size:.75rem;color:#64748b">Estado: {inc['estado'].upper()}</div>

                </div>

                """, unsafe_allow_html=True)



                if inc["estado"] == "pendiente":

                    with st.expander(f"Justificar incidencia {inc['fecha']}"):

                        with st.form(f"just_{inc['id']}"):

                            resolucion = st.text_area("Justificación / explicación")

                            if st.form_submit_button("Enviar justificación"):

                                db.resolve_incident(inc["id"], resolucion, user["id"])

                                st.success("Incidencia justificada.")

                                st.rerun()



# ── Page: Mi Equipo (Manager) ─────────────────────────────────────────────────



def page_equipo():

    if not is_role("admin", "manager"): st.error("🔒 Acceso no autorizado."); return

    user = current_user()

    hoy = date.today()

    st.title("👥 Mi Equipo")



    team = db.get_team(user["id"])

    if not team:

        st.info("No tienes empleados asignados.")

        return



    # Today's overview

    st.subheader(f"Estado hoy — {hoy.strftime('%d/%m/%Y')}")

    cols = st.columns(len(team))

    for i, emp in enumerate(team):

        estado = db.get_day_status(emp["id"], hoy, emp.get("comunidad_autonoma","madrid"))

        entries = db.get_day_entries(emp["id"], hoy)

        worked = db.calc_worked_hours(entries)

        with cols[i]:

            st.markdown(f"""

            <div class="card" style="text-align:center">

                <div style="font-weight:600">{emp['nombre']}</div>

                <div style="font-size:.8rem;color:#64748b">{emp['apellidos']}</div>

                <div style="margin:8px 0">{status_badge(estado)}</div>

                <div style="font-size:.9rem"><b>{fmt_h(worked)}</b></div>

            </div>

            """, unsafe_allow_html=True)



    # Team incidents

    st.subheader("Incidencias del equipo")

    incidents = db.get_team_incidents(user["id"], estado="pendiente")

    if not incidents:

        st.success("✅ Sin incidencias pendientes en el equipo.")

    else:

        for inc in incidents:

            with st.expander(f"{inc['emp_nombre']} — {inc['tipo'].replace('_',' ').title()} ({inc['fecha']})"):

                st.write(inc["descripcion"])

                with st.form(f"resolve_inc_{inc['id']}"):

                    res = st.text_input("Resolución")

                    if st.form_submit_button("Resolver"):

                        db.resolve_incident(inc["id"], res, user["id"])

                        st.success("Resuelta."); st.rerun()



    # Weekly hours per employee

    st.subheader("Horas esta semana por empleado")

    lunes = hoy - timedelta(days=hoy.weekday())

    team_ids = [e["id"] for e in team]

    entries_sem = db.get_all_entries_range(lunes, hoy, user_ids=team_ids)

    if entries_sem:

        df = pd.DataFrame(entries_sem)

        result = df.groupby(["emp_nombre","fecha"]).apply(

            lambda g: db.calc_worked_hours(g.to_dict("records"))

        ).reset_index(name="horas")

        fig = px.bar(result, x="fecha", y="horas", color="emp_nombre",

                     title="Horas semanales por empleado",

                     labels={"fecha":"Fecha","horas":"Horas","emp_nombre":"Empleado"},

                     barmode="group")

        fig.add_hline(y=8, line_dash="dash", line_color="#6366f1")

        fig.add_hline(y=9, line_dash="dash", line_color="#ef4444")

        st.plotly_chart(fig, use_container_width=True)



# ── Page: Aprobar Solicitudes ─────────────────────────────────────────────────



def page_aprobar():

    if not is_role("admin", "manager"): st.error("🔒 Acceso no autorizado."); return

    user = current_user()

    st.title("✅ Aprobar Solicitudes")



    pending = db.get_pending_requests(user["id"] if not is_role("admin") else None)

    if not pending:

        st.success("✅ No hay solicitudes pendientes.")

        return



    st.info(f"**{len(pending)}** solicitudes pendientes de revisión.")



    for r in pending:

        with st.expander(

            f"**{r['emp_nombre']}** · {r['tipo'].replace('_',' ').title()} · "

            f"{r['fecha_inicio']} → {r['fecha_fin']} ({r['dias_laborables']}d laborables)"

        ):

            if r["comentario_empleado"]:

                st.write(f"Comentario: {r['comentario_empleado']}")

            st.caption(f"Solicitado: {r['fecha_solicitud'][:10]}")

            comentario_mgr = st.text_input("Comentario al empleado (opcional)",

                                           key=f"cm_{r['id']}")

            c1, c2 = st.columns(2)

            if c1.button("✅ Aprobar", key=f"apr_{r['id']}", type="primary"):

                nuevo_estado = "aprobada"

                db.resolve_request(r["id"], nuevo_estado, user["id"], comentario_mgr)

                db.add_notification(r["user_id"], "vacaciones_aprobada",

                    "Tu solicitud ha sido aprobada",

                    f"{r['tipo'].replace('_',' ').title()}: {r['fecha_inicio']} → {r['fecha_fin']}")

                db.audit(user["id"], "aprobar_solicitud", "vacation_requests", r["id"])

                # Send email notification

                try:

                    emp = db.get_user_by_id(r["user_id"])

                    if emp and emp.get("email"):

                        bal = db.get_vacation_balance(emp["id"], date.today().year)

                        used = bal.get("used", 0) if bal else 0

                        anuales = emp.get("dias_vacaciones_anuales", 22)

                        db.send_vacation_email(

                            emp["email"], f"{emp['nombre']} {emp['apellidos']}",

                            r.get("tipo","vacaciones"), nuevo_estado,

                            r["fecha_inicio"], r["fecha_fin"],

                            r.get("dias_laborables",0),

                            comentario_mgr, used, anuales

                        )

                except Exception:

                    pass

                st.success("Aprobada."); st.rerun()

            if c2.button("❌ Denegar", key=f"den_{r['id']}"):

                nuevo_estado = "denegada"

                db.resolve_request(r["id"], nuevo_estado, user["id"], comentario_mgr)

                db.add_notification(r["user_id"], "vacaciones_denegada",

                    "Tu solicitud ha sido denegada",

                    f"{r['tipo'].replace('_',' ').title()}: {r['fecha_inicio']} → {r['fecha_fin']}"

                    + (f"\nMotivo: {comentario_mgr}" if comentario_mgr else ""))

                db.audit(user["id"], "denegar_solicitud", "vacation_requests", r["id"])

                # Send email notification

                try:

                    emp = db.get_user_by_id(r["user_id"])

                    if emp and emp.get("email"):

                        bal = db.get_vacation_balance(emp["id"], date.today().year)

                        used = bal.get("used", 0) if bal else 0

                        anuales = emp.get("dias_vacaciones_anuales", 22)

                        db.send_vacation_email(

                            emp["email"], f"{emp['nombre']} {emp['apellidos']}",

                            r.get("tipo","vacaciones"), nuevo_estado,

                            r["fecha_inicio"], r["fecha_fin"],

                            r.get("dias_laborables",0),

                            comentario_mgr, used, anuales

                        )

                except Exception:

                    pass

                st.warning("Denegada."); st.rerun()



# ── Page: Dashboard Admin ─────────────────────────────────────────────────────



def page_admin_dashboard():

    if not is_role("admin"): st.error("🔒 Acceso no autorizado."); return

    st.title("📊 Dashboard Administrador")

    hoy = date.today()

    lunes = hoy - timedelta(days=hoy.weekday())

    inicio_mes = date(hoy.year, hoy.month, 1)



    stats = db.get_global_stats()



    c1, c2, c3, c4 = st.columns(4)

    c1.metric("👥 Empleados activos", stats["total_users"])

    c2.metric("✅ Fichajes hoy", stats["total_entries_today"])

    c3.metric("📋 Solicitudes pendientes", stats["pending_requests"])

    c4.metric("⚠️ Incidencias abiertas", stats["open_incidents"])



    # Compliance semáforo

    users = db.get_all_users(activos_only=True)

    green = yellow = red = 0

    for u in users:

        s = db.get_day_status(u["id"], hoy, u.get("comunidad_autonoma","madrid"))

        if s == "verde": green += 1

        elif s == "amarillo": yellow += 1

        elif s == "rojo": red += 1



    total_hoy = max(1, green + yellow + red)

    st.subheader("🚦 Semáforo de cumplimiento — hoy")

    col_g, col_y, col_r = st.columns(3)

    col_g.markdown(f"""<div class="kpi"><div class="kpi-val" style="color:#10b981">{green}</div>

        <div class="kpi-label">✓ Cumplimiento correcto</div>

        <div style="font-size:.8rem;color:#94a3b8">{int(green/total_hoy*100)}%</div></div>""",

        unsafe_allow_html=True)

    col_y.markdown(f"""<div class="kpi"><div class="kpi-val" style="color:#f59e0b">{yellow}</div>

        <div class="kpi-label">⚠ Riesgo medio</div>

        <div style="font-size:.8rem;color:#94a3b8">{int(yellow/total_hoy*100)}%</div></div>""",

        unsafe_allow_html=True)

    col_r.markdown(f"""<div class="kpi"><div class="kpi-val" style="color:#ef4444">{red}</div>

        <div class="kpi-label">✗ Riesgo alto</div>

        <div style="font-size:.8rem;color:#94a3b8">{int(red/total_hoy*100)}%</div></div>""",

        unsafe_allow_html=True)



    # Weekly chart all users

    st.subheader("Horas trabajadas por empleado — semana actual")

    entries_sem = db.get_all_entries_range(lunes, hoy)

    if entries_sem:

        df = pd.DataFrame(entries_sem)

        result = df.groupby(["emp_nombre","fecha"]).apply(

            lambda g: db.calc_worked_hours(g.to_dict("records"))

        ).reset_index(name="horas")

        fig = px.bar(result, x="fecha", y="horas", color="emp_nombre",

                     barmode="group", title="Horas semanales (todos los empleados)",

                     labels={"fecha":"Fecha","horas":"Horas","emp_nombre":"Empleado"})

        fig.add_hline(y=8, line_dash="dash", line_color="#6366f1", annotation_text="8h")

        fig.add_hline(y=9, line_dash="dash", line_color="#ef4444", annotation_text="9h máx.")

        st.plotly_chart(fig, use_container_width=True)

    else:

        st.info("Sin registros esta semana.")



    # Per-user vacation balance

    st.subheader("Balance de vacaciones — todos los empleados")

    bal_data = []

    for u in users:

        bal = db.get_vacation_balance(u["id"], hoy.year)

        bal_data.append({

            "Empleado": f"{u['nombre']} {u['apellidos']}",

            "Total días": bal["total"],

            "Disfrutados": bal["used"],

            "Pendiente aprobación": bal["pending"],

            "Disponibles": bal["remaining"],

        })

    if bal_data:

        df_bal = pd.DataFrame(bal_data)

        fig = px.bar(df_bal, x="Empleado", y=["Disfrutados","Pendiente aprobación","Disponibles"],

                     barmode="stack", title=f"Vacaciones {hoy.year} — mín. 30 días naturales (Art.38 ET)",

                     color_discrete_map={"Disfrutados":"#4f46e5","Pendiente aprobación":"#f59e0b","Disponibles":"#e2e8f0"})

        st.plotly_chart(fig, use_container_width=True)



# ── Page: Gestión Usuarios ────────────────────────────────────────────────────



# -- Page: Gestion Usuarios ---------------------------------------------------

ROLE_LABEL = {"admin":"Director","manager":"Director de Area","empleado":"Empleado"}
ROLE_ICON  = {"admin":"crown","manager":"folder","empleado":"person"}
ROLE_EMOJI = {"admin":"👑","manager":"🗂️","empleado":"👤"}
CARGO_BY_ROLE = {
    "admin":   ["Director"],
    "manager": ["Director de Area"],
    "empleado":["Comercial","Administrativo","Otro"],
}

def _mgr_candidates(role, all_users):
    if role == "admin":   return []
    if role == "manager": return [u for u in all_users if u["role"] == "admin"]
    return [u for u in all_users if u["role"] in ("manager","admin")]

def _build_tree(users):
    rows = []
    seen = set()
    def add_subtree(u, lvl):
        if u["id"] in seen: return
        seen.add(u["id"])
        rows.append((lvl, u))
        for c in users:
            if c.get("manager_id") == u["id"] and c["id"] not in seen:
                add_subtree(c, lvl + 1)
    for u in users:
        if u["role"] == "admin":
            add_subtree(u, 0)
    for u in users:
        if u["role"] == "manager" and u["id"] not in seen:
            add_subtree(u, 1)
    for u in users:
        if u["id"] not in seen:
            rows.append((2, u))
    return rows

def page_usuarios():
    if not is_role("admin"): st.error("🔒 Acceso no autorizado."); return
    st.title("Gestión de Usuarios")
    tab_tree, tab_edit, tab_new = st.tabs(["Organigrama", "Editar usuario", "Nuevo usuario"])

    all_users = db.get_all_users()
    depts     = db.get_departments()
    dept_opts = {"Sin departamento": None}
    dept_opts.update({d["nombre"]: d["id"] for d in depts})

    with tab_tree:
        st.markdown("#### Estructura jerárquica")
        tree = _build_tree(all_users)
        if not tree:
            st.info("No hay usuarios.")
        else:
            for lvl, u in tree:
                pad   = u"\u00a0" * (lvl * 8)
                icon  = ROLE_EMOJI.get(u["role"], "👤")
                cargo = u.get("cargo") or ROLE_LABEL.get(u["role"], u["role"])
                mgr   = u.get("manager_nombre") or "—"
                dept  = u.get("dept_nombre") or "—"
                estado= "Activo" if u.get("activo", 1) else "Inactivo"
                arrows = ["", "L-- ", "    L-- "]
                arrow  = arrows[min(lvl, 2)]
                st.markdown(
                    pad + arrow + icon + " **" + u["nombre"] + " " + u["apellidos"] + "**"
                    + "  `" + cargo + "`"
                    + "  Depto: " + dept
                    + "  Reporta a: " + mgr
                    + "  " + estado,
                    unsafe_allow_html=False
                )

    with tab_edit:
        if not all_users:
            st.info("Sin usuarios.")
            return

        opts = {
            ROLE_EMOJI.get(u["role"],"") + " " + u["nombre"] + " " + u["apellidos"] + " · " + ROLE_LABEL.get(u["role"],u["role"]): u["id"]
            for u in all_users
        }
        sel = st.selectbox("Seleccionar usuario", list(opts.keys()), key="edit_sel")
        uid = opts[sel]
        ud  = next((u for u in all_users if u["id"] == uid), {})
        roles = ["empleado","manager","admin"]
        comunidades = list(COMUNIDADES_MAP.keys())

        with st.form("edit_user"):
            st.markdown("**Datos personales**")
            c1, c2 = st.columns(2)
            nombre    = c1.text_input("Nombre",    value=ud.get("nombre",""))
            apellidos = c2.text_input("Apellidos", value=ud.get("apellidos",""))
            email     = c1.text_input("Email",     value=ud.get("email",""))
            telefono  = c2.text_input("Teléfono",  value=ud.get("telefono",""))

            st.markdown("**Puesto y jerarquía**")
            c3, c4 = st.columns(2)
            role = c3.selectbox(
                "Nivel jerárquico", roles,
                format_func=lambda x: ROLE_EMOJI[x] + " " + ROLE_LABEL[x],
                index=roles.index(ud.get("role","empleado")) if ud.get("role") in roles else 0,
            )
            cargo_opts = CARGO_BY_ROLE.get(ud.get("role","empleado"), ["Comercial","Administrativo","Otro"])
            cur_cargo  = ud.get("cargo") or cargo_opts[0]
            cargo = c4.selectbox(
                "Cargo", cargo_opts,
                index=cargo_opts.index(cur_cargo) if cur_cargo in cargo_opts else 0,
            )

            st.markdown("**Organización**")
            c5, c6 = st.columns(2)
            dept_keys     = list(dept_opts.keys())
            cur_dept_name = ud.get("dept_nombre") or "Sin departamento"
            dept_idx      = dept_keys.index(cur_dept_name) if cur_dept_name in dept_keys else 0
            dept_sel      = c5.selectbox("Departamento", dept_keys, index=dept_idx)

            mgr_cands = _mgr_candidates(ud.get("role","empleado"), all_users)
            mgr_opts_edit = {"Sin responsable": None}
            mgr_opts_edit.update({
                ROLE_EMOJI.get(u["role"],"") + " " + u["nombre"] + " " + u["apellidos"] + " · " + ROLE_LABEL.get(u["role"],u["role"]): u["id"]
                for u in mgr_cands if u["id"] != uid
            })
            cur_mgr = ud.get("manager_id")
            mgr_idx = list(mgr_opts_edit.values()).index(cur_mgr) if cur_mgr in mgr_opts_edit.values() else 0
            mgr_sel = c6.selectbox("Responsable directo", list(mgr_opts_edit.keys()), index=mgr_idx)

            c7, c8 = st.columns(2)
            horas    = c7.number_input("Horas semanales", value=float(ud.get("horas_semanales",40)), min_value=1.0, max_value=40.0, step=0.5)
            dias_vac = c8.number_input("Días vacaciones/año", value=int(ud.get("dias_vacaciones_anuales",22)), min_value=1, max_value=60)
            comunidad = c7.selectbox(
                "Comunidad autónoma", comunidades,
                index=comunidades.index(ud.get("comunidad_autonoma","madrid")) if ud.get("comunidad_autonoma") in comunidades else 0,
                format_func=lambda x: COMUNIDADES_MAP[x],
            )
            activo = c8.checkbox("Activo", value=bool(ud.get("activo",1)))

            st.markdown("**Ubicación**")
            c9, c10 = st.columns(2)
            prov_list = [""] + PROVINCIAS_ES
            cur_prov  = ud.get("provincia","") or ""
            provincia = c9.selectbox("Provincia", prov_list,
                                     index=prov_list.index(cur_prov) if cur_prov in prov_list else 0)
            localidad = c10.text_input("Localidad", value=ud.get("localidad","") or "")
            comunidad_final = PROVINCIA_COMUNIDAD.get(provincia, comunidad) if provincia else comunidad
            new_pw = st.text_input("Nueva contraseña (vacío = sin cambios)", type="password")

            if st.form_submit_button("Guardar cambios", type="primary"):
                kw = dict(
                    nombre=nombre, apellidos=apellidos, email=email, telefono=telefono,
                    role=role, cargo=cargo,
                    department_id=dept_opts[dept_sel],
                    manager_id=mgr_opts_edit[mgr_sel],
                    horas_semanales=horas, dias_vacaciones_anuales=dias_vac,
                    comunidad_autonoma=comunidad_final, activo=int(activo),
                    provincia=provincia, localidad=localidad,
                )
                if new_pw:
                    kw["password"] = new_pw
                db.update_user(uid, **kw)
                db.audit(current_user()["id"], "editar_usuario", "users", uid)
                # Si el admin se editó a sí mismo, actualizar la sesión
                if uid == current_user()["id"]:
                    updated_user = db.get_user_by_id(uid)
                    if updated_user:
                        updated_safe = {k: v for k, v in updated_user.items() if k != "password_hash"}
                        st.session_state["user"] = updated_safe
                st.success("Usuario actualizado.")
                st.rerun()

    with tab_new:
        with st.form("new_user"):
            st.markdown("**Datos personales**")
            c1, c2 = st.columns(2)
            username  = c1.text_input("Usuario (login) *")
            password  = c2.text_input("Contraseña *", type="password")
            nombre    = c1.text_input("Nombre *")
            apellidos = c2.text_input("Apellidos *")
            email     = c1.text_input("Email")
            telefono  = c2.text_input("Teléfono")

            st.markdown("**Puesto y jerarquía**")
            c3, c4 = st.columns(2)
            role_n  = c3.selectbox(
                "Nivel jerárquico", ["empleado","manager","admin"],
                format_func=lambda x: ROLE_EMOJI[x] + " " + ROLE_LABEL[x],
                key="new_role",
            )
            cargo_n = c4.selectbox(
                "Cargo", ["Comercial","Administrativo","Director de Area","Director","Otro"],
                key="new_cargo",
            )

            st.markdown("**Organización**")
            c5, c6 = st.columns(2)
            dept_n = c5.selectbox("Departamento", list(dept_opts.keys()), key="new_dept")
            all_mgr_cands = [u for u in all_users if u["role"] in ("manager","admin")]
            mgr_opts_new = {"Sin responsable": None}
            mgr_opts_new.update({
                ROLE_EMOJI.get(u["role"],"") + " " + u["nombre"] + " " + u["apellidos"] + " · " + ROLE_LABEL.get(u["role"],u["role"]): u["id"]
                for u in all_mgr_cands
            })
            mgr_n = c6.selectbox("Responsable directo", list(mgr_opts_new.keys()), key="new_mgr")

            c7, c8 = st.columns(2)
            horas_n    = c7.number_input("Horas semanales", value=40.0, min_value=1.0, max_value=40.0, step=0.5, key="new_h")
            dias_vac_n = c8.number_input("Días vacaciones/año", value=22, min_value=1, max_value=60, key="new_dv")
            comunidad_n = c7.selectbox("Comunidad autónoma", list(COMUNIDADES_MAP.keys()),
                                       format_func=lambda x: COMUNIDADES_MAP[x], key="new_com")

            st.markdown("**Ubicación**")
            c9, c10 = st.columns(2)
            prov_n = c9.selectbox("Provincia", [""]+PROVINCIAS_ES, key="new_prov")
            loc_n  = c10.text_input("Localidad", key="new_loc")

            if st.form_submit_button("Crear usuario", type="primary"):
                if not username or not password or not nombre or not apellidos:
                    st.error("Campos obligatorios: usuario, contraseña, nombre, apellidos.")
                else:
                    try:
                        com_final = PROVINCIA_COMUNIDAD.get(prov_n, comunidad_n) if prov_n else comunidad_n
                        uid_new = db.create_user(
                            username, password, nombre, apellidos, email,
                            role_n, dept_opts[dept_n], mgr_opts_new[mgr_n], com_final, horas_n, dias_vac_n
                        )
                        db.update_user(uid_new, cargo=cargo_n, provincia=prov_n, localidad=loc_n, telefono=telefono)
                        db.audit(current_user()["id"], "crear_usuario", "users", uid_new)
                        st.success("Usuario " + username + " creado.")
                        st.rerun()
                    except Exception as ex:
                        st.error("Error: " + str(ex))

        st.markdown("---")
        st.subheader("Departamentos")
        for d in db.get_departments():
            st.write("- " + d["nombre"])
        with st.form("new_dept"):
            dept_nombre = st.text_input("Nuevo departamento")
            if st.form_submit_button("Añadir"):
                if dept_nombre:
                    db.create_department(dept_nombre)
                    st.success("Departamento creado.")
                    st.rerun()


def page_festivos():

    if not is_role("admin"): st.error("🔒 Acceso no autorizado."); return

    st.title("📅 Gestión de Festivos")

    año = st.number_input("Año", value=date.today().year, min_value=2020, max_value=2035)



    comunidad_filter = st.selectbox("Comunidad autónoma",

                                    ["todas"] + list(COMUNIDADES_MAP.keys()),

                                    format_func=lambda x: "Todas" if x=="todas" else COMUNIDADES_MAP[x])



    festivos = db.get_holidays_df(comunidad=None if comunidad_filter=="todas" else comunidad_filter,

                                  año=año)

    if festivos:

        df = pd.DataFrame(festivos)

        st.dataframe(

            df[["fecha","descripcion","tipo","comunidad_autonoma"]].rename(columns={

                "fecha":"Fecha","descripcion":"Descripción",

                "tipo":"Tipo","comunidad_autonoma":"Comunidad"}),

            use_container_width=True, hide_index=True

        )

        # Delete option

        fid_opts = {f"{f['fecha']} — {f['descripcion']}": f["id"] for f in festivos}

        del_sel = st.selectbox("Eliminar festivo", list(fid_opts.keys()))

        if st.button("🗑️ Eliminar seleccionado"):

            db.delete_holiday(fid_opts[del_sel])

            db.audit(current_user()["id"], "eliminar_festivo")

            st.success("Eliminado."); st.rerun()



    st.subheader("➕ Añadir festivo")

    with st.form("add_fest"):

        c1, c2, c3, c4 = st.columns(4)

        fecha_f = c1.date_input("Fecha")

        desc_f = c2.text_input("Descripción")

        tipo_f = c3.selectbox("Tipo", ["nacional","autonomico","local"])

        com_f = c4.selectbox("Comunidad", ["todas"] + list(COMUNIDADES_MAP.keys()),

                             format_func=lambda x: "Todas" if x=="todas" else COMUNIDADES_MAP[x])

        if st.form_submit_button("Añadir festivo"):

            db.add_holiday(fecha_f, desc_f, tipo_f, com_f, fecha_f.year)

            db.audit(current_user()["id"], "añadir_festivo")

            st.success("Festivo añadido."); st.rerun()



# ── Page: Exportaciones ───────────────────────────────────────────────────────



def page_exportaciones():

    if not is_role("admin"): st.error("🔒 Acceso no autorizado."); return

    st.title("📥 Exportaciones")

    st.markdown("""<div class="legal-box">

    <b>RDL 8/2019 — Art. 34.9 ET</b>: Registros conservados <b>4 años</b> y disponibles

    para trabajadores, sus representantes y la Inspección de Trabajo y Seguridad Social.

    </div>""", unsafe_allow_html=True)



    hoy = date.today()

    users = db.get_all_users(activos_only=True)

    emp_opts = {"Todos los empleados": None}

    emp_opts.update({f"{u['nombre']} {u['apellidos']}": u["id"] for u in users})



    c1, c2, c3, c4 = st.columns(4)

    emp_sel = c1.selectbox("Empleado", list(emp_opts.keys()))

    emp_id = emp_opts[emp_sel]

    f_ini = c2.date_input("Desde", value=date(hoy.year, hoy.month, 1))

    f_fin = c3.date_input("Hasta", value=hoy)

    tipo_exp = c4.selectbox("Informe", [

        "Registro horario completo",

        "Resumen mensual",

        "Vacaciones y ausencias",

        "Horas extra",

        "Incidencias",

    ])



    if st.button("📊 Generar Excel", type="primary"):

        out = BytesIO()

        with pd.ExcelWriter(out, engine="openpyxl") as writer:

            if tipo_exp == "Registro horario completo":

                if emp_id:

                    entries = db.get_entries_range(emp_id, f_ini, f_fin)

                else:

                    entries = db.get_all_entries_range(f_ini, f_fin)

                if entries:

                    df = pd.DataFrame(entries)

                    # Añadir columna de horas trabajadas diarias
                    if not df.empty and "tipo" in df.columns:
                        try:
                            df["hora_dt"] = pd.to_datetime(df["hora"], errors="coerce")
                            entradas = df[df["tipo"]=="entrada"].groupby("fecha")["hora_dt"].min()
                            salidas = df[df["tipo"]=="salida"].groupby("fecha")["hora_dt"].max()
                            horas_dia = ((salidas - entradas).dt.total_seconds() / 3600).round(2)
                            df["horas_trabajadas"] = df["fecha"].map(horas_dia)
                        except Exception:
                            df["horas_trabajadas"] = ""

                    df = df[["fecha","emp_nombre","tipo","hora","horas_trabajadas","observaciones","is_manual","ip"]].rename(columns={

                        "fecha":"Fecha","emp_nombre":"Empleado","tipo":"Tipo","hora":"Hora",

                        "horas_trabajadas":"Horas trabajadas","observaciones":"Observaciones","is_manual":"Manual","ip":"IP"})

                    df.to_excel(writer, sheet_name="Registro Horario", index=False)



            elif tipo_exp == "Resumen mensual":

                if emp_id:

                    entries = db.get_entries_range(emp_id, f_ini, f_fin)

                else:

                    entries = db.get_all_entries_range(f_ini, f_fin)

                if entries:

                    df = pd.DataFrame(entries)

                    def calc(g):

                        return db.calc_worked_hours(g.to_dict("records"))

                    res = df.groupby(["emp_nombre","fecha"]).apply(calc).reset_index(name="horas")

                    res["mes"] = pd.to_datetime(res["fecha"]).dt.to_period("M").astype(str)

                    summary = res.groupby(["mes","emp_nombre"]).agg(

                        dias=("fecha","nunique"), horas_totales=("horas","sum"),

                        max_dia=("horas","max")).reset_index()

                    summary["horas_extra"] = (summary["horas_totales"] - summary["dias"]*8).clip(lower=0)

                    summary.columns = ["Mes","Empleado","Días","Horas totales","Máx/día","Horas extra aprox."]

                    summary.to_excel(writer, sheet_name="Resumen Mensual", index=False)



            elif tipo_exp == "Vacaciones y ausencias":

                reqs = db.get_all_requests()

                if reqs:

                    df = pd.DataFrame(reqs)

                    df[["emp_nombre","tipo","fecha_inicio","fecha_fin","dias_laborables",

                        "estado","comentario_empleado","comentario_manager","manager_nombre","fecha_solicitud"]].rename(columns={

                        "emp_nombre":"Empleado","tipo":"Tipo","fecha_inicio":"Desde",

                        "fecha_fin":"Hasta","dias_laborables":"Días lab.","estado":"Estado",

                        "comentario_empleado":"Comentario emp.","comentario_manager":"Respuesta mgr.",

                        "manager_nombre":"Aprobado por","fecha_solicitud":"Fecha solicitud"

                    }).to_excel(writer, sheet_name="Vacaciones", index=False)



            elif tipo_exp == "Horas extra":

                if emp_id:

                    entries = db.get_entries_range(emp_id, f_ini, f_fin)

                else:

                    entries = db.get_all_entries_range(f_ini, f_fin)

                if entries:

                    df = pd.DataFrame(entries)

                    def calc(g):

                        return db.calc_worked_hours(g.to_dict("records"))

                    res = df.groupby(["emp_nombre","fecha"]).apply(calc).reset_index(name="horas")

                    res["extra"] = (res["horas"] - 8).clip(lower=0)

                    res["alerta"] = res["horas"].apply(lambda h: "SUPERA 9H — Art.34.3 ET" if h > 9 else "")

                    summary = res.groupby("emp_nombre").agg(

                        total_extra=("extra","sum"), dias_con_extra=("extra",lambda x:(x>0).sum())

                    ).reset_index()

                    summary["limite_legal"] = 80

                    summary["excede_Art35"] = summary["total_extra"] > 80

                    summary.columns = ["Empleado","Total H.Extra","Días con extra","Límite legal/año","Excede Art.35 ET"]

                    summary.to_excel(writer, sheet_name="Resumen Horas Extra", index=False)

                    res[["emp_nombre","fecha","horas","extra","alerta"]].rename(columns={

                        "emp_nombre":"Empleado","fecha":"Fecha","horas":"Horas totales",

                        "extra":"Horas extra","alerta":"Alerta legal"

                    }).to_excel(writer, sheet_name="Detalle diario", index=False)



            elif tipo_exp == "Incidencias":

                incs = db.get_all_incidents()

                if incs:

                    df = pd.DataFrame(incs)

                    df[["emp_nombre","fecha","tipo","descripcion","estado","resolucion","created_at"]].rename(columns={

                        "emp_nombre":"Empleado","fecha":"Fecha","tipo":"Tipo",

                        "descripcion":"Descripción","estado":"Estado",

                        "resolucion":"Resolución","created_at":"Creada"

                    }).to_excel(writer, sheet_name="Incidencias", index=False)



            # Legal metadata

            pd.DataFrame({

                "Campo":["Generado el","Período","Normativa","Obligación conservación"],

                "Valor":[

                    datetime.now().strftime("%d/%m/%Y %H:%M"),

                    f"{f_ini:%d/%m/%Y} — {f_fin:%d/%m/%Y}",

                    "RDL 8/2019 · ET Arts. 34, 35, 38 · RGPD",

                    "4 años (Art. 34.9 ET)",

                ]

            }).to_excel(writer, sheet_name="Info Legal", index=False)

            # ── Professional Excel formatting ─────────────────────────────────
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            def _format_sheet(ws, title_color="4F46E5"):
                header_fill = PatternFill("solid", fgColor=title_color)
                header_font = Font(bold=True, color="FFFFFF", size=10)
                thin = Side(style="thin", color="CCCCCC")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border
                ws.freeze_panes = "A2"
                for col_idx, col in enumerate(ws.columns, 1):
                    max_len = max((len(str(cell.value or "")) for cell in col), default=8)
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)
                ws.row_dimensions[1].height = 20

            for sheet_name in writer.sheets:
                _format_sheet(writer.sheets[sheet_name])



        out.seek(0)

        fname = f"ficha_{tipo_exp.lower().replace(' ','_')}_{f_ini:%Y%m%d}_{f_fin:%Y%m%d}.xlsx"

        st.download_button("⬇️ Descargar Excel", data=out, file_name=fname,

                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

                           type="primary")

        st.success(f"✅ Archivo generado: **{fname}**")

        db.audit(current_user()["id"], "exportacion", datos={"tipo": tipo_exp, "f_ini": str(f_ini), "f_fin": str(f_fin)})



# ── Page: Auditoría ───────────────────────────────────────────────────────────



def page_auditoria():

    if not is_role("admin"): st.error("🔒 Acceso no autorizado."); return

    st.title("🔍 Auditoría y Trazabilidad")

    st.markdown("""<div class="legal-box">

    <b>RGPD + RDL 8/2019</b>: Historial completo e inmutable. Ningún registro puede eliminarse físicamente.

    Solo correcciones auditadas con trazabilidad completa.

    </div>""", unsafe_allow_html=True)



    logs = db.get_audit_logs(limit=200)

    if not logs:

        st.info("Sin registros de auditoría.")

        return



    df = pd.DataFrame(logs)

    st.dataframe(

        df[["created_at","user_nombre","accion","tabla","registro_id","datos","ip"]].rename(columns={

            "created_at":"Timestamp","user_nombre":"Usuario","accion":"Acción",

            "tabla":"Tabla","registro_id":"ID Registro","datos":"Datos","ip":"IP"

        }),

        use_container_width=True, hide_index=True

    )



    # Export audit log

    out = BytesIO()

    df.to_excel(out, index=False)

    out.seek(0)

    st.download_button("⬇️ Exportar log auditoría", data=out,

                       file_name=f"auditoria_{date.today()}.xlsx",

                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



# ── Email one-click approval page ─────────────────────────────────────────────



def page_token_action(token: str):
    """Handles one-click approve/deny links from manager notification emails."""
    st.markdown(LOGIN_CSS, unsafe_allow_html=True)

    _, col, _ = st.columns([1, 1.2, 1])

    with col:

        st.markdown("""
<div class="odk-brand-wrap">
  <svg width="48" height="52" viewBox="0 0 56 60" fill="none" xmlns="http://www.w3.org/2000/svg" style="display:block;margin:0 auto">
    <path d="M3 5 L28 39 L53 5 Z" stroke="#CA8A04" stroke-width="1.9" fill="rgba(202,138,4,0.08)" stroke-linejoin="round"/>
    <path d="M17 21 L28 39 L39 21 Z" fill="rgba(202,138,4,0.18)"/>
    <line x1="28" y1="39" x2="28" y2="53" stroke="#CA8A04" stroke-width="1.9" stroke-linecap="round"/>
    <line x1="15" y1="53" x2="41" y2="53" stroke="#CA8A04" stroke-width="2.5" stroke-linecap="round"/>
  </svg>
  <span class="odk-brand-name">ODK</span>
  <div class="odk-gold-line"></div>
  <span class="odk-brand-sub">Gestión de solicitudes</span>
</div>
""", unsafe_allow_html=True)

        result = db.validate_and_use_token(token)
        status = result.get("status")

        if status == "ok":
            aprobada = result["action"] == "approve"
            color  = "#16a34a" if aprobada else "#dc2626"
            bg     = "rgba(22,163,74,.10)" if aprobada else "rgba(220,38,38,.10)"
            icon   = "✓" if aprobada else "✗"
            titulo = "Solicitud aprobada" if aprobada else "Solicitud denegada"
            st.markdown(f"""
<div style="background:{bg};border:1px solid {color};border-radius:14px;padding:28px 24px;text-align:center;margin-bottom:16px;">
  <div style="font-size:2.8rem;color:{color};margin-bottom:8px;">{icon}</div>
  <div style="color:#f0e8d5;font-family:'Montserrat',sans-serif;font-size:1.1rem;font-weight:700;letter-spacing:.06em;margin-bottom:6px;">{titulo}</div>
  <div style="color:rgba(220,195,140,.70);font-family:'Montserrat',sans-serif;font-size:.85rem;">
    {result.get('emp_nombre','')} &middot; {result.get('tipo','').replace('_',' ').title()}<br>
    {result.get('fecha_ini','')} &rarr; {result.get('fecha_fin','')}
  </div>
</div>
<p style="color:rgba(220,195,140,.45);font-family:'Montserrat',sans-serif;font-size:.75rem;text-align:center;letter-spacing:.04em;">
  Se ha notificado al empleado por email. Puedes cerrar esta ventana.
</p>
""", unsafe_allow_html=True)

        elif status == "used":
            already = result.get("estado", result.get("action",""))
            st.markdown(f"""
<div style="background:rgba(202,138,4,.08);border:1px solid rgba(202,138,4,.25);border-radius:14px;padding:24px;text-align:center;">
  <div style="color:#CA8A04;font-size:1.8rem;margin-bottom:8px;">⚠</div>
  <div style="color:#f0e8d5;font-family:'Montserrat',sans-serif;font-weight:700;margin-bottom:6px;">Enlace ya utilizado</div>
  <div style="color:rgba(220,195,140,.60);font-family:'Montserrat',sans-serif;font-size:.82rem;">Esta solicitud ya fue procesada{f' ({already})' if already else ''}.</div>
</div>
""", unsafe_allow_html=True)

        elif status == "expired":
            st.markdown("""
<div style="background:rgba(220,38,38,.08);border:1px solid rgba(220,38,38,.22);border-radius:14px;padding:24px;text-align:center;">
  <div style="color:#dc2626;font-size:1.8rem;margin-bottom:8px;">⏰</div>
  <div style="color:#f0e8d5;font-family:'Montserrat',sans-serif;font-weight:700;margin-bottom:6px;">Enlace caducado</div>
  <div style="color:rgba(220,195,140,.60);font-family:'Montserrat',sans-serif;font-size:.82rem;">Los enlaces de aprobación caducan a las 72 h. Accede a la app para gestionar la solicitud.</div>
</div>
""", unsafe_allow_html=True)

        else:
            st.markdown("""
<div style="background:rgba(220,38,38,.08);border:1px solid rgba(220,38,38,.22);border-radius:14px;padding:24px;text-align:center;">
  <div style="color:#dc2626;font-size:1.8rem;margin-bottom:8px;">✗</div>
  <div style="color:#f0e8d5;font-family:'Montserrat',sans-serif;font-weight:700;margin-bottom:6px;">Enlace no válido</div>
  <div style="color:rgba(220,195,140,.60);font-family:'Montserrat',sans-serif;font-size:.82rem;">Si crees que esto es un error, contacta con el administrador.</div>
</div>
""", unsafe_allow_html=True)



# ── Main ──────────────────────────────────────────────────────────────────────



def main():

    if not st.session_state.get("_db_ready"):
        db.init_db()
        st.session_state["_db_ready"] = True

    st.markdown(STYLES, unsafe_allow_html=True)

    # ── One-click email token handler (no login required) ─────────────────────
    tok = st.query_params.get("tok", "")
    if tok:
        page_token_action(tok)
        return

    if not logged_in():

        page_login()

        return



    page = render_sidebar()



    # Clean page name (strip emoji prefix)

    clean = page.strip()



    if "Perfil" in clean:

        page_perfil()

    elif "Fichaje" in clean:

        page_fichaje()

    elif "Calendario" in clean:

        page_calendario()

    elif "Vacaciones" in clean:

        page_vacaciones()

    elif "Notificaciones" in clean:

        page_notificaciones()

    elif "Equipo" in clean:

        page_equipo()

    elif "Aprobar" in clean:

        page_aprobar()

    elif "Dashboard" in clean:

        page_admin_dashboard()

    elif "Usuarios" in clean:

        page_usuarios()

    elif "Festivos" in clean:

        page_festivos()

    elif "Exportaciones" in clean:

        page_exportaciones()

    elif "Auditoría" in clean:

        page_auditoria()

    else:

        page_fichaje()



if __name__ == "__main__":

    main()

